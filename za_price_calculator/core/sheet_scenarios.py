"""Построение листа 'Сценарный анализ' - сводная таблица по 5 сценариям и сигналы."""
from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.worksheet import Worksheet

from za_price_calculator.config import PALETTE, SHEETS
from za_price_calculator.styling.styles import (
    HDR_FILL,
    HDR_FONT,
    NUM2,
    PCT,
    align,
    border_thin,
    fill,
    font,
)


def build_scenarios_sheet(ws: Worksheet, n_rows: int) -> None:
    """
    Строит сводный лист сравнения сценариев ценообразования и блок автосигналов.

    :param ws: целевой лист.
    :param n_rows: количество строк данных на листе 'Расчеты' (N).
    """
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE.purple

    for col, w in (("A", 3), ("B", 32), ("C", 24), ("D", 16), ("E", 16), ("F", 16), ("G", 16), ("H", 30)):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:H2")
    t = ws["B2"]
    t.value = "СЦЕНАРНЫЙ АНАЛИЗ - Сравнение 5 сценариев по всем позициям"
    t.font = font(bold=True, size=14, color=PALETTE.white)
    t.fill = fill(PALETTE.dark_blue)
    t.alignment = align("center")
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:H3")
    s = ws["B3"]
    s.value = ("Кол-во продаж вводится вручную на листе Расчеты (кол. T) или на листе Продажи. "
               "При отсутствии - расчёты по ценам активны.")
    s.font = font(size=9, color="595959")
    s.alignment = align("left", indent=1)
    ws.row_dimensions[3].height = 16

    if n_rows == 0:
        ws.merge_cells("B5:H6")
        c = ws["B5"]
        c.value = "Нет данных для сценарного анализа - загрузите прайс-лист с товарами."
        c.font = font(bold=True, size=12, color=PALETTE.accent_red)
        c.alignment = align("center", wrap=True)
        return

    cs = f"'{SHEETS.calculations}'"
    last = n_rows + 2

    hdr_row = 5
    for ci, h in enumerate(
        ["Сценарий", "Описание", "Ср.новая цена", "Ср.новая маржа", "Ср.ВП/ед", "Сумм.прогн.ВП", "Вывод"], start=2
    ):
        c = ws.cell(hdr_row, ci, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = align("center", wrap=True)
        c.border = border_thin()
    ws.row_dimensions[hdr_row].height = 22

    def avg_if(col):
        return f'=IFERROR(AVERAGE(IF({cs}!{col}3:{col}{last}<>"",{cs}!{col}3:{col}{last})),"")'

    def avg_if_gt0(col):
        return f'=IFERROR(AVERAGE(IF({cs}!{col}3:{col}{last}>0,{cs}!{col}3:{col}{last})),"")'

    scen_rows = [
        (6, "Текущая", "Действующая цена ЗЯ",
         avg_if("D"), avg_if("P"), avg_if("Q"), f"=SUM({cs}!V3:V{last})",
         "Базовый уровень", fill(PALETTE.grey_bg)),
        (7, "С1: Средняя рынка", "Цена = средняя по конкурентам",
         avg_if("X"), avg_if("Z"), avg_if("AA"), f"=SUM({cs}!BI3:BI{last})",
         "Выравнивание по рынку", fill("EBF3FB")),
        (8, "С2: Медиана рынка", "Цена = медианная по конкурентам",
         avg_if("AE"), avg_if("AG"), avg_if("AH"), f"=SUM({cs}!BK3:BK{last})",
         "Устойчив к выбросам", fill("EBF3FB")),
        (9, "С3: Мин.конкурент", "Цена = минимальный конкурент",
         avg_if("AL"), avg_if("AN"), avg_if("AO"), f"=SUM({cs}!BM3:BM{last})",
         "Риск потери маржи", fill(PALETTE.red_bg)),
        (10, "С4: Произв.цена", "Пользовательская цена (кол.AS в Расчетах)",
         avg_if_gt0("AS"), avg_if("AU"), avg_if("AV"), f"=SUM({cs}!BO3:BO{last})",
         "Пользовательский сценарий", fill("F0E6FF")),
        (11, "С5: Целевая маржа", "Цена по целевой марже (кол.AZ в Расчетах)",
         avg_if("BA"), avg_if("BC"), avg_if("BD"), f"=SUM({cs}!BQ3:BQ{last})",
         "Маржинальный сценарий", fill(PALETTE.green_bg)),
    ]
    for (row, lbl, desc, f_price, f_mrg, f_vp, f_sum, note, rfill) in scen_rows:
        ws.row_dimensions[row].height = 22
        for ci, val in enumerate([lbl, desc, f_price, f_mrg, f_vp, f_sum, note], start=2):
            c = ws.cell(row, ci, val)
            c.fill = rfill
            c.border = border_thin()
            c.font = font(bold=(ci == 2), size=10)
            c.alignment = align("center") if ci > 3 else align("left", indent=1)
            if ci == 4:
                c.number_format = NUM2
            elif ci == 5:
                c.number_format = PCT
            elif ci in (6, 7):
                c.number_format = NUM2

    ws.row_dimensions[13].height = 8
    ws.merge_cells("B14:H14")
    c = ws["B14"]
    c.value = "АВТОМАТИЧЕСКИЕ СИГНАЛЫ"
    c.font = font(bold=True, size=11, color=PALETTE.white)
    c.fill = fill(PALETTE.dark_blue)
    c.alignment = align("center")
    ws.row_dimensions[14].height = 22

    signals = [
        ("Позиций с ценой ВЫШЕ рынка", f'=COUNTIF({cs}!E3:E10000,"Выше рынка")',
         "Рассмотреть снижение цены до С1 или С2"),
        ("Позиций с ценой НИЖЕ рынка", f'=COUNTIF({cs}!E3:E10000,"Ниже рынка")',
         "Потенциал повышения цены без потери конкурентоспособности"),
        ("Позиций с КРИТИЧНЫМ риском ВП (С3)", f'=COUNTIF({cs}!S3:S10000,"КРИТИЧНО")',
         "Снижение до мин.конкурента даёт маржу <5% - опасно"),
        ("Позиций с УМЕРЕННЫМ риском (С3)", f'=COUNTIF({cs}!S3:S10000,"УМЕРЕННЫЙ")',
         "Требует мониторинга при изменении цены"),
        ("Позиций с ХОРОШЕЙ маржой (>20%)", f'=COUNTIF({cs}!R3:R10000,"Хорошая(>20%)")',
         "Высокая ценовая гибкость"),
        ("Позиций с НИЗКОЙ маржой (<10%)", f'=COUNTIF({cs}!R3:R10000,"Низкая(<10%)")',
         "Приоритет: переговоры с поставщиком / ценовая корректировка"),
    ]
    for j, (lbl, frm, rec) in enumerate(signals):
        row_j = 15 + j
        ws.row_dimensions[row_j].height = 22
        ws.merge_cells(f"B{row_j}:C{row_j}")
        c = ws.cell(row_j, 2, lbl)
        c.font = font(bold=True, size=10)
        c.fill = fill(PALETTE.grey_bg)
        c.alignment = align("left", indent=1)
        c.border = border_thin()

        c = ws.cell(row_j, 4, frm)
        c.font = font(bold=True, size=14)
        c.number_format = "#,##0"
        c.alignment = align("center")
        c.fill = fill(PALETTE.light_blue)
        c.border = border_thin()

        ws.merge_cells(f"E{row_j}:H{row_j}")
        c = ws.cell(row_j, 5, rec)
        c.font = font(size=9, color="595959")
        c.alignment = align("left", indent=1, wrap=True)
        c.fill = fill(PALETTE.grey_bg)
        c.border = border_thin()

    ws.conditional_formatting.add(
        "D15:D20", CellIsRule(operator="greaterThan", formula=["0"], fill=fill(PALETTE.red_bg))
    )
