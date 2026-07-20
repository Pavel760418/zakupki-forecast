"""Построение листа 'Исходные данные' из загруженного DataFrame."""
from __future__ import annotations

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from za_price_calculator.config import SOURCE_COLUMNS
from za_price_calculator.styling.styles import (
    HDR_ALIGN,
    HDR_FILL,
    HDR_FONT,
    INPUT_FONT,
    NUM2,
    align,
    border_thin,
    font,
)

_HEADER_LABELS = SOURCE_COLUMNS
_WIDTHS = [42, 18, 16, 18, 16, 16, 14, 12, 14, 18, 20]


def build_source_sheet(ws: Worksheet, df: pd.DataFrame) -> int:
    """
    Заполняет лист 'Исходные данные' и оформляет его как Excel Table.

    :param ws: целевой лист.
    :param df: нормализованный DataFrame с колонками SOURCE_COLUMNS.
    :return: количество строк данных (N).
    """
    ws.sheet_view.showGridLines = False

    n = len(df)
    for ci, (header, width) in enumerate(zip(_HEADER_LABELS, _WIDTHS), start=1):
        c = ws.cell(1, ci, header)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = HDR_ALIGN
        c.border = border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22

    for ri, row in df.iterrows():
        er = ri + 2
        for ci, col in enumerate(SOURCE_COLUMNS, start=1):
            val = row[col]
            cell = ws.cell(er, ci, val if pd.notna(val) and val != "" else None)
            cell.border = border_thin()
            if ci == 1:
                cell.font = font(size=10)
                cell.alignment = align("left", indent=1)
            elif ci == 2:
                cell.font = font(size=10, color="595959")
                cell.alignment = align("center")
            else:
                cell.font = INPUT_FONT
                cell.alignment = align("right")
                cell.number_format = NUM2

    last_row = n + 1
    if n > 0:
        table = Table(displayName="tbl_Src", ref=f"A1:{get_column_letter(len(SOURCE_COLUMNS))}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True, showFirstColumn=True
        )
        ws.add_table(table)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SOURCE_COLUMNS))}{last_row}"
    return n
