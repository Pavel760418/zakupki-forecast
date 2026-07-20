"""Построение листа 'Инструкция'."""
from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from za_price_calculator.config import PALETTE, SHEETS
from za_price_calculator.styling.styles import align, fill, font


def _title(ws: Worksheet, row: int, text: str) -> None:
    ws.merge_cells(f"B{row}:K{row}")
    c = ws.cell(row, 2, text)
    c.font = font(bold=True, size=14, color=PALETTE.white)
    c.fill = fill(PALETTE.dark_blue)
    c.alignment = align("left", indent=1)
    ws.row_dimensions[row].height = 30


def _section(ws: Worksheet, row: int, text: str) -> None:
    ws.merge_cells(f"B{row}:K{row}")
    c = ws.cell(row, 2, text)
    c.font = font(bold=True, size=10, color=PALETTE.white)
    c.fill = fill(PALETTE.mid_blue)
    c.alignment = align("left", indent=1)
    ws.row_dimensions[row].height = 20


def _line(ws: Worksheet, row: int, text: str, bold: bool = False) -> None:
    ws.merge_cells(f"B{row}:K{row}")
    c = ws.cell(row, 2, text)
    c.font = font(bold=bold, size=10)
    c.fill = fill(PALETTE.grey_bg)
    c.alignment = align("left", wrap=True, indent=1)
    ws.row_dimensions[row].height = 18


def build_instructions_sheet(ws: Worksheet) -> None:
    """Строит лист инструкции с описанием листов, ручного ввода, формул и навигацией."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE.dark_blue
    ws.column_dimensions["A"].width = 3
    for ci in range(2, 12):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    _title(ws, 2, "\U0001F34F  ЗЕЛЕНОЕ ЯБЛОКО - Калькулятор цен, маржи и сценариев v1.0")
    ws.row_dimensions[3].height = 6

    _section(ws, 4, "НАЗНАЧЕНИЕ ФАЙЛА")
    _line(ws, 5, "Инструмент для коммерческого блока: анализ цены, наценки, маржи, сравнение с "
                 "рынком и сценарное моделирование по всем позициям.")

    _section(ws, 7, "ЛИСТЫ ФАЙЛА")
    lines = [
        f"{SHEETS.source}  - товарная матрица: цены ЗЯ и конкурентов. Загружается из исходного файла.",
        f"{SHEETS.sales}          - продажи по штрихкоду (кол-во, выручка, ВП). Опционально.",
        f"{SHEETS.calculations}          - все автоматические формулы. Ручной ввод только в синих ячейках.",
        f"{SHEETS.dashboard}        - KPI-карточки, светофоры, топы. Обновляется автоматически.",
        f"{SHEETS.scenarios} - сводная таблица по 5 сценариям и автоматические рекомендации.",
    ]
    for j, t in enumerate(lines, start=8):
        _line(ws, j, t)

    _section(ws, 14, "ПОЛЯ РУЧНОГО ВВОДА (синий цвет = вводить вручную)")
    input_lines = [
        f"Лист «{SHEETS.sales}»: Штрихкод, Кол-во продаж, Выручка, Валовая прибыль.",
        "Лист «Расчеты», колонка T (Кол-во продаж) - можно вводить прямо здесь.",
        "Лист «Расчеты», колонка AS (С4: Новая цена вручную) - произвольная цена для сценария 4.",
        "Лист «Расчеты», колонка AZ (С5: Целевая маржа) - например 0.25 для целевой маржи 25%.",
    ]
    for j, t in enumerate(input_lines, start=15):
        _line(ws, j, t)

    _section(ws, 20, "ФОРМУЛЫ - КЛЮЧЕВАЯ ЛОГИКА")
    formula_lines = [
        "Наценка  = (Розн. цена - Закуп. цена) / Закуп. цена",
        "Маржа    = (Розн. цена - Закуп. цена) / Розн. цена",
        "ВП/ед    = Розн. цена - Закуп. цена",
        "Цена по целевой марже (С5): Цена = Закупочная цена / (1 - Целевая маржа)",
        "Необходимый рост продаж = Текущая суммарная ВП / ВП/ед нового сценария - Текущие продажи",
    ]
    for j, t in enumerate(formula_lines, start=21):
        _line(ws, j, t)

    _section(ws, 27, "ЦВЕТОВЫЕ ОБОЗНАЧЕНИЯ")
    color_lines = [
        ("Синий текст", "Поле ручного ввода - вводить вручную"),
        ("Черный текст", "Формула - не изменять"),
        ("Зеленый фон", "Цена ниже рынка / хорошая маржа"),
        ("Красный фон", "Цена выше рынка / риск потери маржи"),
        ("Желтый фон", "Нейтральная зона / требует внимания"),
    ]
    for j, (k, v) in enumerate(color_lines, start=28):
        ws.merge_cells(f"B{j}:D{j}")
        ws.merge_cells(f"E{j}:K{j}")
        c1 = ws.cell(j, 2, k)
        c2 = ws.cell(j, 5, v)
        c1.font = font(bold=True, size=10)
        c1.fill = fill(PALETTE.grey_bg)
        c1.alignment = align("left", indent=1)
        c2.font = font(size=10)
        c2.fill = fill(PALETTE.grey_bg)
        c2.alignment = align("left", indent=1)
        ws.row_dimensions[j].height = 18

    _section(ws, 34, "НАВИГАЦИЯ")
    nav_links = [
        (SHEETS.source, f"'{SHEETS.source}'!A1"),
        (SHEETS.sales, f"'{SHEETS.sales}'!A1"),
        (SHEETS.calculations, f"'{SHEETS.calculations}'!A1"),
        (SHEETS.dashboard, f"'{SHEETS.dashboard}'!B2"),
        (SHEETS.scenarios, f"'{SHEETS.scenarios}'!B2"),
    ]
    for j, (label, loc) in enumerate(nav_links, start=35):
        ws.merge_cells(f"B{j}:K{j}")
        c = ws.cell(j, 2, label)
        c.hyperlink = Hyperlink(ref=c.coordinate, location=loc)
        from openpyxl.styles import Font as _Font
        c.font = _Font(name="Calibri", bold=True, size=11, color="0070C0", underline="single")
        c.fill = fill(PALETTE.light_blue)
        c.alignment = align("left", indent=2)
        ws.row_dimensions[j].height = 20
