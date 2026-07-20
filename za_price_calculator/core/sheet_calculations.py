"""
Построение листа 'Расчеты' - ядро калькулятора.
Содержит все формулы: анализ рынка, текущую экономику, 5 сценариев ценообразования
и моделирование продаж. Формулы идентичны оригинальному ZA_Price_Calculator.xlsx.
"""
from __future__ import annotations

from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from za_price_calculator.config import CALC_GROUPS, PALETTE, SHEETS
from za_price_calculator.styling.styles import (
    NUM0,
    NUM2,
    PCT,
    align,
    border_thin,
    fill,
    font,
)

INPUT_FONT = font(size=10, color=PALETTE.input_text)
FORMULA_FONT = font(size=10, color=PALETTE.formula_text)
XREF_FONT = font(size=10, color=PALETTE.xref_text)
INPUT_FILL = fill(PALETTE.input_blue_bg)

COLUMN_DEFS: list[tuple[int, str, int, str, bool]] = [
    (1, "Наименование", 40, "@", False),
    (2, "Штрихкод", 16, "@", False),
    (3, "Закупочная цена", 14, NUM2, False),
    (4, "Розничная цена ЗЯ", 14, NUM2, False),
    (5, "Сигнал цены", 16, "@", False),
    (6, "Ср.цена рынка", 14, NUM2, False),
    (7, "Медиана рынка", 14, NUM2, False),
    (8, "Мин.конкурент", 14, NUM2, False),
    (9, "Макс.конкурент", 14, NUM2, False),
    (10, "Откл.от средней", 13, PCT, False),
    (11, "Откл.от медианы", 13, PCT, False),
    (12, "Откл.от минимума", 13, PCT, False),
    (13, "Рейтинг цены", 11, "0", False),
    (14, "Позиц.на рынке", 14, "@", False),
    (15, "Наценка тек.%", 13, PCT, False),
    (16, "Маржа тек.%", 13, PCT, False),
    (17, "ВП/ед тек.", 13, NUM2, False),
    (18, "Сигнал маржи", 14, "@", False),
    (19, "Риск потери ВП", 14, "@", False),
    (20, "Кол-во продаж", 14, NUM0, True),
    (21, "Выручка тек.", 14, NUM2, False),
    (22, "ВП тек.суммарная", 14, NUM2, False),
    (23, "Маржа тек.общая%", 14, PCT, False),
    (24, "С1:Цена", 12, NUM2, False),
    (25, "С1:Наценка%", 11, PCT, False),
    (26, "С1:Маржа%", 11, PCT, False),
    (27, "С1:ВП/ед", 12, NUM2, False),
    (28, "С1:Дельта цены%", 11, PCT, False),
    (29, "С1:Дельта маржи%", 11, PCT, False),
    (30, "С1:Дельта ВП/ед", 12, NUM2, False),
    (31, "С2:Цена", 12, NUM2, False),
    (32, "С2:Наценка%", 11, PCT, False),
    (33, "С2:Маржа%", 11, PCT, False),
    (34, "С2:ВП/ед", 12, NUM2, False),
    (35, "С2:Дельта цены%", 11, PCT, False),
    (36, "С2:Дельта маржи%", 11, PCT, False),
    (37, "С2:Дельта ВП/ед", 12, NUM2, False),
    (38, "С3:Цена", 12, NUM2, False),
    (39, "С3:Наценка%", 11, PCT, False),
    (40, "С3:Маржа%", 11, PCT, False),
    (41, "С3:ВП/ед", 12, NUM2, False),
    (42, "С3:Дельта цены%", 11, PCT, False),
    (43, "С3:Дельта маржи%", 11, PCT, False),
    (44, "С3:Дельта ВП/ед", 12, NUM2, False),
    (45, "С4:Нов.цена[ВВОД]", 14, NUM2, True),
    (46, "С4:Наценка%", 11, PCT, False),
    (47, "С4:Маржа%", 11, PCT, False),
    (48, "С4:ВП/ед", 12, NUM2, False),
    (49, "С4:Дельта цены%", 11, PCT, False),
    (50, "С4:Дельта маржи%", 11, PCT, False),
    (51, "С4:Дельта ВП/ед", 12, NUM2, False),
    (52, "С5:Цел.маржа[ВВОД]", 14, PCT, True),
    (53, "С5:Цена", 12, NUM2, False),
    (54, "С5:Наценка%", 11, PCT, False),
    (55, "С5:Маржа%", 11, PCT, False),
    (56, "С5:ВП/ед", 12, NUM2, False),
    (57, "С5:Дельта цены%", 11, PCT, False),
    (58, "С5:Дельта маржи%", 11, PCT, False),
    (59, "С5:Дельта ВП/ед", 12, NUM2, False),
    (60, "С1:Прогн.выручка", 14, NUM2, False),
    (61, "С1:Прогн.ВП", 14, NUM2, False),
    (62, "С2:Прогн.выручка", 14, NUM2, False),
    (63, "С2:Прогн.ВП", 14, NUM2, False),
    (64, "С3:Прогн.выручка", 14, NUM2, False),
    (65, "С3:Прогн.ВП", 14, NUM2, False),
    (66, "С4:Прогн.выручка", 14, NUM2, False),
    (67, "С4:Прогн.ВП", 14, NUM2, False),
    (68, "С5:Прогн.выручка", 14, NUM2, False),
    (69, "С5:Прогн.ВП", 14, NUM2, False),
    (70, "С1:Треб.рост шт", 13, NUM0, False),
    (71, "С1:Треб.рост %", 12, PCT, False),
    (72, "С3:Треб.рост шт", 13, NUM0, False),
    (73, "С3:Треб.рост %", 12, PCT, False),
]

LAST_COL = 73


def _write(ws: Worksheet, row: int, col: int, formula, num_fmt: str, is_input: bool = False) -> None:
    cell = ws.cell(row, col, formula)
    cell.border = border_thin()
    cell.number_format = num_fmt
    if col == 1:
        cell.alignment = align("left", indent=1)
    elif col in (2, 5, 14, 18, 19):
        cell.alignment = align("center")
    else:
        cell.alignment = align("right")
    if is_input:
        cell.font = INPUT_FONT
        cell.fill = INPUT_FILL
    elif col in (1, 2, 3, 4):
        cell.font = XREF_FONT
    else:
        cell.font = FORMULA_FONT


def build_calculations_sheet(ws: Worksheet, n_rows: int) -> None:
    """
    Строит лист 'Расчеты' с полным набором формул для n_rows строк товаров.

    :param ws: целевой лист openpyxl.
    :param n_rows: количество строк данных на листе 'Исходные данные' (N).
    """
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE.orange

    col_grp_fill: dict[int, str] = {}
    for (c1, c2, label, hexcolor) in CALC_GROUPS:
        for ci in range(c1, c2 + 1):
            col_grp_fill[ci] = hexcolor
        if c2 > c1:
            ws.merge_cells(f"{get_column_letter(c1)}1:{get_column_letter(c2)}1")
        cell = ws.cell(1, c1, label)
        cell.fill = fill(hexcolor)
        cell.font = font(bold=True, size=9, color=PALETTE.white)
        cell.alignment = align("center")
        cell.border = border_thin()
    ws.row_dimensions[1].height = 18

    for (ci, header, width, num_fmt, is_input) in COLUMN_DEFS:
        cell = ws.cell(2, ci, header)
        cell.fill = fill(col_grp_fill.get(ci, PALETTE.dark_blue))
        cell.font = font(bold=True, size=9, color=PALETTE.white)
        cell.alignment = align("center", wrap=True)
        cell.border = border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 36

    src = f"'{SHEETS.source}'"

    if n_rows == 0:
        return

    for ri in range(n_rows):
        r = ri + 3
        sr = ri + 2

        _write(ws, r, 1, f"={src}!A{sr}", "@")
        _write(ws, r, 2, f"={src}!B{sr}", "@")
        _write(ws, r, 3, f"={src}!C{sr}", NUM2)
        _write(ws, r, 4, f"={src}!D{sr}", NUM2)

        _write(ws, r, 5,
               f'=IF(OR(D{r}="",F{r}=""),"-",'
               f'IF(D{r}>F{r}*1.05,"Выше рынка",'
               f'IF(D{r}<F{r}*0.95,"Ниже рынка","На уровне")))', "@")

        comp = ",".join(f"{src}!{get_column_letter(ci)}{sr}" for ci in range(5, 12))
        _write(ws, r, 6, f'=IFERROR(AVERAGE({comp}),"")', NUM2)
        _write(ws, r, 7, f'=IFERROR(MEDIAN({comp}),"")', NUM2)
        _write(ws, r, 8, f'=IFERROR(MIN({comp}),"")', NUM2)
        _write(ws, r, 9, f'=IFERROR(MAX({comp}),"")', NUM2)

        for ci, ref in ((10, "F"), (11, "G"), (12, "H")):
            _write(ws, r, ci, f'=IF(AND(D{r}<>"",{ref}{r}<>""),(D{r}-{ref}{r})/{ref}{r},"")', PCT)

        _write(ws, r, 13, f'=IFERROR(RANK(D{r},({comp})),"")', "0")
        _write(ws, r, 14,
               f'=IF(M{r}="","-",IF(M{r}=1,"Дороже всех",'
               f'IF(M{r}=2,"2-й по цене",IF(M{r}<=4,"Средний","Ниже конкурентов"))))', "@")

        _write(ws, r, 15, f'=IF(AND(C{r}>0,D{r}>0),(D{r}-C{r})/C{r},"")', PCT)
        _write(ws, r, 16, f'=IF(AND(C{r}>0,D{r}>0),(D{r}-C{r})/D{r},"")', PCT)
        _write(ws, r, 17, f'=IF(AND(C{r}>0,D{r}>0),D{r}-C{r},"")', NUM2)

        _write(ws, r, 18,
               f'=IF(P{r}="","-",IF(P{r}<0.1,"Низкая(<10%)",'
               f'IF(P{r}<0.2,"Средняя(10-20%)","Хорошая(>20%)")))', "@")
        _write(ws, r, 19,
               f'=IF(OR(H{r}="",C{r}=""),"-",IF((H{r}-C{r})/H{r}<0.05,"КРИТИЧНО",'
               f'IF((H{r}-C{r})/H{r}<0.15,"УМЕРЕННЫЙ","НИЗКИЙ")))', "@")

        _write(ws, r, 20, None, NUM0, is_input=True)
        _write(ws, r, 21, f'=IF(T{r}>0,T{r}*D{r},"")', NUM2)
        _write(ws, r, 22, f'=IF(T{r}>0,T{r}*Q{r},"")', NUM2)
        _write(ws, r, 23, f'=IF(U{r}>0,V{r}/U{r},"")', PCT)

        def scenario_block(price_col, mkup_col, mrg_col, vp_col, dprc_col, dmrg_col, dvp_col, price_formula):
            if price_formula is not None:
                _write(ws, r, price_col, price_formula, NUM2)
            price_L = get_column_letter(price_col)
            mrg_L = get_column_letter(mrg_col)
            vp_L = get_column_letter(vp_col)
            _write(ws, r, mkup_col, f'=IF(AND({price_L}{r}<>"",C{r}>0),({price_L}{r}-C{r})/C{r},"")', PCT)
            _write(ws, r, mrg_col, f'=IF(AND({price_L}{r}<>"",{price_L}{r}>0),({price_L}{r}-C{r})/{price_L}{r},"")', PCT)
            _write(ws, r, vp_col, f'=IF(AND({price_L}{r}<>"",C{r}>0),{price_L}{r}-C{r},"")', NUM2)
            _write(ws, r, dprc_col, f'=IF(AND({price_L}{r}<>"",D{r}>0),({price_L}{r}-D{r})/D{r},"")', PCT)
            _write(ws, r, dmrg_col,
                   f'=IF(AND({mrg_L}{r}<>"",P{r}<>""),({mrg_L}{r}-P{r})/ABS(IF(P{r}=0,1,P{r})),"")', PCT)
            _write(ws, r, dvp_col, f'=IF(AND({vp_L}{r}<>"",Q{r}<>""),{vp_L}{r}-Q{r},"")', NUM2)

        scenario_block(24, 25, 26, 27, 28, 29, 30, f'=IF(F{r}<>"",F{r},"")')
        scenario_block(31, 32, 33, 34, 35, 36, 37, f'=IF(G{r}<>"",G{r},"")')
        scenario_block(38, 39, 40, 41, 42, 43, 44, f'=IF(H{r}<>"",H{r},"")')

        _write(ws, r, 45, None, NUM2, is_input=True)
        for ci, expr, nf in (
            (46, f'=IF(AND(AS{r}>0,C{r}>0),(AS{r}-C{r})/C{r},"")', PCT),
            (47, f'=IF(AND(AS{r}>0,AS{r}>0),(AS{r}-C{r})/AS{r},"")', PCT),
            (48, f'=IF(AND(AS{r}>0,C{r}>0),AS{r}-C{r},"")', NUM2),
            (49, f'=IF(AND(AS{r}>0,D{r}>0),(AS{r}-D{r})/D{r},"")', PCT),
            (50, f'=IF(AND(AU{r}<>"",P{r}<>""),(AU{r}-P{r})/ABS(IF(P{r}=0,1,P{r})),"")', PCT),
            (51, f'=IF(AND(AV{r}<>"",Q{r}<>""),AV{r}-Q{r},"")', NUM2),
        ):
            _write(ws, r, ci, expr, nf)

        _write(ws, r, 52, None, PCT, is_input=True)
        for ci, expr, nf in (
            (53, f'=IF(AND(AZ{r}>0,C{r}>0),C{r}/(1-AZ{r}),"")', NUM2),
            (54, f'=IF(AND(BA{r}<>"",C{r}>0),(BA{r}-C{r})/C{r},"")', PCT),
            (55, f'=IF(BA{r}<>"",AZ{r},"")', PCT),
            (56, f'=IF(AND(BA{r}<>"",C{r}>0),BA{r}-C{r},"")', NUM2),
            (57, f'=IF(AND(BA{r}<>"",D{r}>0),(BA{r}-D{r})/D{r},"")', PCT),
            (58, f'=IF(AND(BC{r}<>"",P{r}<>""),(BC{r}-P{r})/ABS(IF(P{r}=0,1,P{r})),"")', PCT),
            (59, f'=IF(AND(BD{r}<>"",Q{r}<>""),BD{r}-Q{r},"")', NUM2),
        ):
            _write(ws, r, ci, expr, nf)

        for ci, expr, nf in (
            (60, f'=IF(AND(T{r}>0,X{r}<>""),T{r}*X{r},"")', NUM2),
            (61, f'=IF(AND(T{r}>0,AA{r}<>""),T{r}*AA{r},"")', NUM2),
            (62, f'=IF(AND(T{r}>0,AE{r}<>""),T{r}*AE{r},"")', NUM2),
            (63, f'=IF(AND(T{r}>0,AH{r}<>""),T{r}*AH{r},"")', NUM2),
            (64, f'=IF(AND(T{r}>0,AL{r}<>""),T{r}*AL{r},"")', NUM2),
            (65, f'=IF(AND(T{r}>0,AO{r}<>""),T{r}*AO{r},"")', NUM2),
            (66, f'=IF(AND(T{r}>0,AS{r}>0),T{r}*AS{r},"")', NUM2),
            (67, f'=IF(AND(T{r}>0,AV{r}<>""),T{r}*AV{r},"")', NUM2),
            (68, f'=IF(AND(T{r}>0,BA{r}<>""),T{r}*BA{r},"")', NUM2),
            (69, f'=IF(AND(T{r}>0,BD{r}<>""),T{r}*BD{r},"")', NUM2),
            (70, f'=IF(AND(T{r}>0,V{r}>0,AA{r}<>0),V{r}/AA{r}-T{r},"")', NUM0),
            (71, f'=IF(AND(BR{r}<>"",T{r}>0),BR{r}/T{r},"")', PCT),
            (72, f'=IF(AND(T{r}>0,V{r}>0,AO{r}<>0),V{r}/AO{r}-T{r},"")', NUM0),
            (73, f'=IF(AND(BT{r}<>"",T{r}>0),BT{r}/T{r},"")', PCT),
        ):
            _write(ws, r, ci, expr, nf)

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(LAST_COL)}{n_rows + 2}"

    last_data_row = n_rows + 2
    ws.conditional_formatting.add(
        f"J3:J{last_data_row}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                        mid_type="num", mid_value=0, mid_color="FFEB84",
                        end_type="max", end_color="F8696B"),
    )
    ws.conditional_formatting.add(
        f"P3:P{last_data_row}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="63BE7B"),
    )
    ws.conditional_formatting.add(
        f"Q3:Q{last_data_row}",
        DataBarRule(start_type="min", end_type="max", color="4472C4"),
    )
