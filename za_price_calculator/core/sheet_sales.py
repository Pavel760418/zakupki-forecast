"""Построение листа 'Продажи' (опциональные данные)."""
from __future__ import annotations

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from za_price_calculator.config import PALETTE, SALES_COLUMNS
from za_price_calculator.styling.styles import (
    HDR_ALIGN,
    HDR_FILL,
    HDR_FONT,
    INPUT_FONT,
    NUM2,
    align,
    border_thin,
    font,
)

_WIDTHS = [20, 16, 16, 18]
_MIN_EMPTY_ROWS = 10


def build_sales_sheet(ws: Worksheet, df) -> None:
    """
    Заполняет лист 'Продажи'. Если df is None или пуст - создаёт пустой шаблон
    для последующего ручного ввода (калькулятор продолжает работать без ошибок).
    """
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE.dark_green

    for ci, (header, width) in enumerate(zip(SALES_COLUMNS, _WIDTHS), start=1):
        c = ws.cell(1, ci, header)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = HDR_ALIGN
        c.border = border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22

    has_data = df is not None and not df.empty
    n_rows = len(df) if has_data else 0
    total_rows = max(n_rows, _MIN_EMPTY_ROWS)

    for ri in range(total_rows):
        er = ri + 2
        for ci, col in enumerate(SALES_COLUMNS, start=1):
            value = None
            if has_data and ri < n_rows:
                v = df.iloc[ri][col]
                value = v if pd.notna(v) and v != "" else None
            cell = ws.cell(er, ci, value)
            cell.border = border_thin()
            cell.font = INPUT_FONT
            cell.alignment = align("center")
            if ci >= 2:
                cell.number_format = NUM2

    last_row = total_rows + 1
    table = Table(displayName="tbl_Sales", ref=f"A1:{get_column_letter(len(SALES_COLUMNS))}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = "A2"

    note_row = last_row + 2
    ws.merge_cells(f"A{note_row}:D{note_row}")
    c = ws.cell(note_row, 1,
                "Заполните по штрихкоду. При отсутствии данных - расчёты на листе "
                "Расчеты работают без ошибок.")
    c.font = font(size=9, color="595959")
    c.alignment = align("left", indent=1)
