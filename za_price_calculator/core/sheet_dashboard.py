"""Построение листа 'Dashboard' - KPI-карточки, топы, светофоры."""
from __future__ import annotations

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from za_price_calculator.config import PALETTE, SHEETS
from za_price_calculator.styling.styles import (
    HDR_ALIGN,
    NUM2,
    PCT,
    PCT_SIGNED,
    align,
    border_thin,
    fill,
    font,
)


def _kpi(ws, row, col, label, formula, num_fmt="#,##0", bg=PALETTE.mid_blue):
    lc = ws.cell(row, col, label)
    lc.font = font(bold=True, size=9, color=PALETTE.white)
    lc.fill = fill(bg)
    lc.alignment = align("center")
    lc.border = border_thin()
    ws.row_dimensions[row].height = 22

    vc = ws.cell(row + 1, col, formula)
    vc.font = font(bold=True, size=18, color="1F3864")
    vc.fill = fill(PALETTE.light_blue)
    vc.alignment = align("center")
    vc.number_format = num_fmt
    vc.border = border_thin()
    ws.row_dimensions[row + 1].height = 34


def _section(ws, row, c1, c2, title, hexcolor=PALETTE.dark_blue):
    ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    c = ws.cell(row, c1, title)
    c.font = font(bold=True, size=11, color=PALETTE.white)
    c.fill = fill(hexcolor)
    c.alignment = align("left", indent=1)
    ws.row_dimensions[row].height = 22


def build_dashboard_sheet(ws: Worksheet, n_rows: int) -> None:
    """
    Строит лист 'Dashboard' с KPI-карточками и топ-10 списками.

    :param ws: целевой лист.
    :param n_rows: количество строк данных на листе 'Расчеты' (N).
    """
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE.accent_green

    for col, w in (("A", 3), ("B", 22), ("C", 22), ("D", 22), ("E", 22), ("F", 22), ("G", 22), ("H", 3)):
        ws.column_dimensions[col].width = w

    cs = f"'{SHEETS.calculations}'"
    last = n_rows + 2 if n_rows else 3

    ws.merge_cells("B2:G2")
    t = ws["B2"]
    t.value = "ЗЕЛЕНОЕ ЯБЛОКО - Dashboard цен и маржи"
    t.font = font(bold=True, size=16, color=PALETTE.white)
    t.fill = fill(PALETTE.dark_blue)
    t.alignment = align("center")
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:G3")
    s = ws["B3"]
    s.value = "Основные показатели по всему прайс-листу | Автообновление"
    s.font = font(size=10, color="595959")
    s.alignment = align("center")
    ws.row_dimensions[3].height = 16

    if n_rows == 0:
        ws.merge_cells("B5:G6")
        c = ws["B5"]
        c.value = "Нет данных для расчёта KPI - загрузите прайс-лист с товарами."
        c.font = font(bold=True, size=12, color=PALETTE.accent_red)
        c.alignment = align("center", wrap=True)
        return

    kpis1 = [
        ("Всего позиций", f"=COUNTA({cs}!A3:A10000)", "#,##0", PALETTE.mid_blue),
        ("Выше рынка", f'=COUNTIF({cs}!E3:E10000,"Выше рынка")', "#,##0", PALETTE.brown),
        ("Ниже рынка", f'=COUNTIF({cs}!E3:E10000,"Ниже рынка")', "#,##0", PALETTE.dark_green),
        ("Средняя маржа", f'=IFERROR(AVERAGE(IF({cs}!P3:P{last}<>"",{cs}!P3:P{last})),"")', PCT, PALETTE.dark_blue),
        ("Средняя наценка", f'=IFERROR(AVERAGE(IF({cs}!O3:O{last}<>"",{cs}!O3:O{last})),"")', PCT, PALETTE.dark_blue),
        ("С критич.риском", f'=COUNTIF({cs}!S3:S10000,"КРИТИЧНО")', "#,##0", PALETTE.accent_red),
    ]
    for col_idx, (lbl, frm, nf, bg) in enumerate(kpis1, start=2):
        _kpi(ws, 5, col_idx, lbl, frm, nf, bg)

    ws.row_dimensions[8].height = 8

    kpis2 = [
        ("Ср.ВП/ед (расч.)", f'=IFERROR(AVERAGE(IF({cs}!Q3:Q{last}<>"",{cs}!Q3:Q{last})),"")', NUM2, PALETTE.mid_blue),
        ("Ср.цена ЗЯ", f'=IFERROR(AVERAGE(IF({cs}!D3:D{last}<>"",{cs}!D3:D{last})),"")', NUM2, PALETTE.mid_blue),
        ("Ср.цена рынка", f'=IFERROR(AVERAGE(IF({cs}!F3:F{last}<>"",{cs}!F3:F{last})),"")', NUM2, PALETTE.mid_blue),
        ("Ср.откл.от рынка", f'=IFERROR(AVERAGE(IF({cs}!J3:J{last}<>"",{cs}!J3:J{last})),"")', PCT_SIGNED, PALETTE.dark_blue),
        ("На уровне рынка", f'=COUNTIF({cs}!E3:E10000,"На уровне")', "#,##0", PALETTE.purple),
        ("Хор.маржа(>20%)", f'=COUNTIF({cs}!R3:R10000,"Хорошая(>20%)")', "#,##0", PALETTE.dark_green),
    ]
    for col_idx, (lbl, frm, nf, bg) in enumerate(kpis2, start=2):
        _kpi(ws, 9, col_idx, lbl, frm, nf, bg)

    ws.row_dimensions[12].height = 8

    _section(ws, 13, 2, 4, "Топ-10 позиций по текущей марже %")
    for ci, h in enumerate(["Наименование", "Маржа %", "ВП/ед"], start=2):
        c = ws.cell(14, ci, h)
        c.fill = fill(PALETTE.mid_blue)
        c.font = font(bold=True, size=10, color=PALETTE.white)
        c.alignment = HDR_ALIGN
        c.border = border_thin()
    ws.row_dimensions[14].height = 18

    top_n = min(10, n_rows)
    for rk in range(1, top_n + 1):
        row_i = 14 + rk
        ws.row_dimensions[row_i].height = 17
        c = ws.cell(row_i, 2,
            f'=IFERROR(INDEX({cs}!A$3:A${last},MATCH(LARGE(IF({cs}!P$3:P${last}<>"",'
            f'{cs}!P$3:P${last}),{rk}),{cs}!P$3:P${last},0)),"")')
        c.font = font(size=10)
        c.alignment = align("left", indent=1)
        c.border = border_thin()

        c = ws.cell(row_i, 3,
            f'=IFERROR(LARGE(IF({cs}!P$3:P${last}<>"",{cs}!P$3:P${last}),{rk}),"")')
        c.font = font(bold=True, size=10)
        c.number_format = PCT
        c.alignment = align("center")
        c.border = border_thin()

        c = ws.cell(row_i, 4,
            f'=IFERROR(INDEX({cs}!Q$3:Q${last},MATCH(LARGE(IF({cs}!P$3:P${last}<>"",'
            f'{cs}!P$3:P${last}),{rk}),{cs}!P$3:P${last},0)),"")')
        c.font = font(size=10)
        c.number_format = NUM2
        c.alignment = align("right")
        c.border = border_thin()

    if top_n > 0:
        ws.conditional_formatting.add(
            f"C15:C{14 + top_n}",
            ColorScaleRule(start_type="min", start_color="FFEB84", end_type="max", end_color="63BE7B"),
        )

    _section(ws, 13, 5, 7, "Топ-10 выше рынка (откл.от средней)", PALETTE.brown)
    for ci, h in enumerate(["Наименование", "Откл.%", "Цена ЗЯ"], start=5):
        c = ws.cell(14, ci, h)
        c.fill = fill(PALETTE.brown)
        c.font = font(bold=True, size=10, color=PALETTE.white)
        c.alignment = HDR_ALIGN
        c.border = border_thin()

    for rk in range(1, top_n + 1):
        row_i = 14 + rk
        c = ws.cell(row_i, 5,
            f'=IFERROR(INDEX({cs}!A$3:A${last},MATCH(LARGE(IF({cs}!J$3:J${last}<>"",'
            f'{cs}!J$3:J${last}),{rk}),{cs}!J$3:J${last},0)),"")')
        c.font = font(size=10)
        c.alignment = align("left", indent=1)
        c.border = border_thin()

        c = ws.cell(row_i, 6,
            f'=IFERROR(LARGE(IF({cs}!J$3:J${last}<>"",{cs}!J$3:J${last}),{rk}),"")')
        c.font = font(bold=True, size=10, color="C00000")
        c.number_format = PCT_SIGNED
        c.alignment = align("center")
        c.border = border_thin()

        c = ws.cell(row_i, 7,
            f'=IFERROR(INDEX({cs}!D$3:D${last},MATCH(LARGE(IF({cs}!J$3:J${last}<>"",'
            f'{cs}!J$3:J${last}),{rk}),{cs}!J$3:J${last},0)),"")')
        c.font = font(size=10)
        c.number_format = NUM2
        c.alignment = align("right")
        c.border = border_thin()

    if top_n > 0:
        ws.conditional_formatting.add(
            f"F15:F{14 + top_n}",
            ColorScaleRule(start_type="min", start_color="FFEB84", end_type="max", end_color="F8696B"),
        )
