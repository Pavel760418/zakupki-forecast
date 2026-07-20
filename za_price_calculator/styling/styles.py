"""
Централизованные объекты стилей openpyxl.
Стили создаются один раз и переиспользуются по всему модулю.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from za_price_calculator.config import PALETTE


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold: bool = False, size: int = 11, color: str = "000000", name: str = "Calibri") -> Font:
    return Font(name=name, bold=bold, size=size, color=color)


def align(h: str = "center", v: str = "center", wrap: bool = False, indent: int = 0) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def border_thin() -> Border:
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


HDR_FILL = fill(PALETTE.dark_blue)
HDR_FONT = font(bold=True, size=10, color=PALETTE.white)
HDR_ALIGN = align("center", wrap=True)

INPUT_FONT = font(size=10, color=PALETTE.input_text)
FORMULA_FONT = font(size=10, color=PALETTE.formula_text)
XREF_FONT = font(size=10, color=PALETTE.xref_text)

INPUT_FILL = fill(PALETTE.input_blue_bg)
GREY_FILL = fill(PALETTE.grey_bg)
LIGHT_BLUE_FILL = fill(PALETTE.light_blue)
GREEN_FILL = fill(PALETTE.green_bg)
RED_FILL = fill(PALETTE.red_bg)

BORDER_THIN = border_thin()

NUM2 = "#,##0.00"
NUM0 = "#,##0"
PCT = "0.0%"
PCT_SIGNED = "+0.0%;-0.0%;\u2014"
