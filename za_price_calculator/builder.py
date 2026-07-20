"""
Оркестратор построения полной книги ZA Price Calculator.
Собирает все листы в правильном порядке и валидирует итоговый файл.
"""
from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from za_price_calculator.config import SHEETS
from za_price_calculator.core.sheet_calculations import build_calculations_sheet
from za_price_calculator.core.sheet_dashboard import build_dashboard_sheet
from za_price_calculator.core.sheet_instructions import build_instructions_sheet
from za_price_calculator.core.sheet_sales import build_sales_sheet
from za_price_calculator.core.sheet_scenarios import build_scenarios_sheet
from za_price_calculator.core.sheet_source import build_source_sheet
from za_price_calculator.exceptions import SheetBuildError

logger = logging.getLogger(__name__)


def build_workbook(source_df: pd.DataFrame, sales_df=None) -> Workbook:
    """
    Строит полную книгу ZA Price Calculator из нормализованных DataFrame.

    :param source_df: DataFrame прайс-листа (см. loader.load_source_file).
    :param sales_df: опциональный DataFrame продаж (см. loader.load_sales_file).
    :return: готовый openpyxl.Workbook с формулами и оформлением.
    :raises SheetBuildError: ошибка при построении любого из листов.
    """
    try:
        wb = Workbook()
        ws_instructions = wb.active
        ws_instructions.title = SHEETS.instructions
        build_instructions_sheet(ws_instructions)

        ws_source = wb.create_sheet(SHEETS.source)
        n_rows = build_source_sheet(ws_source, source_df)

        ws_sales = wb.create_sheet(SHEETS.sales)
        build_sales_sheet(ws_sales, sales_df)

        ws_calc = wb.create_sheet(SHEETS.calculations)
        build_calculations_sheet(ws_calc, n_rows)

        ws_dash = wb.create_sheet(SHEETS.dashboard)
        build_dashboard_sheet(ws_dash, n_rows)

        ws_scen = wb.create_sheet(SHEETS.scenarios)
        build_scenarios_sheet(ws_scen, n_rows)

        logger.info("Книга успешно собрана: %d позиций товаров", n_rows)
        return wb
    except Exception as exc:
        raise SheetBuildError(f"Ошибка при построении книги: {exc}") from exc


def save_workbook(wb: Workbook, output_path) -> Path:
    """
    Сохраняет книгу на диск, создавая директорию назначения при необходимости.

    :param wb: собранная книга openpyxl.
    :param output_path: путь для сохранения .xlsx.
    :return: абсолютный Path сохранённого файла.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info("Файл сохранён: %s", path.resolve())
    return path.resolve()
