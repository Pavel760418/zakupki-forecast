"""
Построение итогового Excel-файла с формулами.

Листы:
  00_Инструкция
  01_Настройки
  02_Дашборд
  03_Расчёт_заказа   ← основная рабочая таблица с формулами
  04_ABC
  05_Риск_OOS
  06_Неликвиды
  07_Избыток
  08_Тренды
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config.settings import OUTPUT_DIR, SETTINGS
from data.merge import GRAIN_STORE
from data.store_utils import NETWORK_STORE_LABEL, is_central_warehouse
from excel.formatting import (
    ALIGN_CENTER,
    ALIGN_WRAP,
    FILL_EDIT,
    FILL_FORMULA,
    FILL_HEADER,
    FILL_READONLY,
    FONT_HEADER,
    THIN,
    add_abc_validation,
    add_oos_conditional,
    add_order_highlight,
    autosize_columns,
    style_header_row,
)
from utils.helpers import ensure_output_dir, safe_str, timestamp_filename

logger = logging.getLogger("zakupki_forecast.excel")

# Индексы колонок листа «Расчёт_заказа» (1-based), 4 релиз
# Жёлтые: магазин/поставщик справочные; остаток, продажи, тренд, коэф.строки, ABC, цена, комментарий
COL = {
    "id": 1,
    "store": 2,
    "supplier": 3,
    "sku": 4,
    "name": 5,
    "barcode": 6,
    "uom": 7,
    "quantum": 8,
    "stock": 9,
    "sales": 10,
    "trend": 11,
    "line_coef": 12,
    "abc": 13,
    "avg_daily": 14,
    "forecast": 15,
    "safety": 16,
    "need": 17,
    "raw_order": 18,
    "rec_order": 19,
    "cover": 20,
    "status": 21,
    "risk": 22,
    "light": 23,
    "price": 24,
    "order_sum": 25,
    "note": 26,
}

HEADERS = [
    "№",
    "Магазин",
    "Поставщик",
    "Артикул",
    "Наименование",
    "Штрихкод",
    "Ед.изм.",
    "Квант",
    "Остаток",
    "Продажи за период",
    "Тренд (коэф.)",
    "Коэф. строки",
    "ABC",
    "Среднедневные",
    "Прогноз продаж",
    "Страховой запас",
    "Потребность",
    "Потребность − остаток",
    "Рекомендуемый заказ",
    "Покрытие, дни",
    "Статус",
    "Риск",
    "Светофор",
    "Цена приходная",
    "Сумма заказа",
    "Комментарий закупщика",
]


def _col_letter(key: str) -> str:
    return get_column_letter(COL[key])


def build_workbook(
    df: pd.DataFrame,
    meta: Dict[str, Any],
    output_path: str | Path | None = None,
    supplier_name: str | None = None,
) -> Path:
    """Создаёт итоговый xlsx (4 релиз): расчёт + упрощённый заказ поставщику."""
    resolved_supplier = safe_str(supplier_name).strip() if supplier_name else ""
    if not resolved_supplier:
        resolved_supplier = safe_str(meta.get("supplier_name", "")).strip()
    resolved_supplier = resolved_supplier or None
    grain = meta.get("grain", "network")

    ensure_output_dir()
    if output_path is None:
        output_path = OUTPUT_DIR / timestamp_filename(SETTINGS["workbook_name_prefix"])
    else:
        output_path = Path(output_path)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _build_instruction(wb, supplier_name=resolved_supplier, grain=grain)
    _build_settings(wb, meta)
    _build_dashboard(wb, df, meta)
    _build_main_calc(wb, df, meta)
    _build_abc_sheet(wb, df)
    _build_filtered_sheet(wb, df[df["is_oos_risk"]], "05_Риск_OOS", "Риски out-of-stock")
    _build_filtered_sheet(wb, df[df["is_dead_stock"]], "06_Неликвиды", "Неликвиды и зависшие позиции")
    _build_filtered_sheet(wb, df[df["is_overstock"]], "07_Избыток", "Избыточные / завышенные запасы")
    _build_trends_sheet(wb, df)

    order_view = df[df["recommended_order"] > 0].copy() if "recommended_order" in df.columns else df.copy()
    # Лист 09: по магазинам показываем потребность точек (чтобы магазины были видны).
    # Итог на ЦС = сумма этих строк (supplier_order_qty на дашборде).
    supplier_view = order_view
    matrix_source = order_view
    if grain == GRAIN_STORE:
        retail_mask = ~df["store"].map(is_central_warehouse) if "store" in df.columns else pd.Series([True] * len(df))
        from data.store_utils import UNKNOWN_STORE_LABEL, NETWORK_STORE_LABEL

        retail = df.loc[
            retail_mask
            & ~df["store"].fillna("").isin({"", UNKNOWN_STORE_LABEL, NETWORK_STORE_LABEL})
        ].copy()
        supplier_view = retail[retail["recommended_order"].fillna(0) > 0].copy()
        # Матрица — все точки, где поставщик есть в данных (в т.ч. заказ 0 = запас достаточен)
        matrix_source = retail

    _build_supplier_order_sheet(wb, supplier_view, resolved_supplier, grain=grain)
    if grain == GRAIN_STORE:
        _build_store_matrix(wb, matrix_source)
        transfers = meta.get("transfers")
        _build_transfer_sheet(wb, transfers if isinstance(transfers, pd.DataFrame) else None, meta)

    wb.save(output_path)
    logger.info("Excel сохранён: %s", output_path)
    return output_path


def _ws(wb: Workbook, title: str):
    return wb.create_sheet(title)


def _build_instruction(
    wb: Workbook, supplier_name: str | None = None, grain: str = "network"
) -> None:
    ws = _ws(wb, "00_Инструкция")
    lines = [
        ("ИНСТРУКЦИЯ ДЛЯ ЗАКУПЩИКА", True, 16),
        ("", False, 11),
        ("1. Назначение файла", True, 13),
        (
            "Файл помогает рассчитать прогноз продаж и рекомендуемый заказ на основе остатков и продаж из 1С. "
            "Главный рабочий лист — «03_Расчёт_заказа». Параметры меняйте на «01_Настройки».",
            False,
            11,
        ),
        ("", False, 11),
        ("2. Какие листы для чего", True, 13),
        ("00_Инструкция — эта памятка.", False, 11),
        ("01_Настройки — периоды, коэффициенты, пороги ABC/OOS/неликвидов. Меняйте жёлтые ячейки.", False, 11),
        ("02_Дашборд — сводные показатели и приоритеты.", False, 11),
        ("03_Расчёт_заказа — основная таблица: формулы пересчитывают заказ при изменении параметров.", False, 11),
        ("04_ABC — классификация товаров A/B/C.", False, 11),
        ("05_Риск_OOS — позиции с риском дефицита.", False, 11),
        ("06_Неликвиды — нет продаж / зависший остаток.", False, 11),
        ("07_Избыток — завышенные запасы.", False, 11),
        ("08_Тренды — рост / спад спроса.", False, 11),
        (
            "09_Заказ_поставщику — упрощённая заявка: наименование, количество, штрихкод, цена, сумма. Есть всегда.",
            False,
            11,
        ),
        (
            "10_Матрица_заказ — шахматка SKU × магазин (только если включена детализация по магазинам).",
            False,
            11,
        ),
        (
            "11_Перемещение_со_склада — что переместить с «Склад основной» в точки "
            "до заказа поставщику. Приоритет: Флагман, далее магазины по объёму продаж.",
            False,
            11,
        ),
        (
            "Четвёртый релиз: в расчёте видны магазин, поставщик, штрихкод и цена. "
            f"Детализация этой книги: {'по магазинам' if grain == GRAIN_STORE else 'сводно по сети'}.",
            False,
            11,
        ),
    ]
    if supplier_name and safe_str(supplier_name).strip():
        lines.append(
            (
                f"В этой книге выбран поставщик: {safe_str(supplier_name).strip()}.",
                False,
                11,
            )
        )
    lines.extend(
        [
        ("", False, 11),
        ("3. Что МОЖНО редактировать (жёлтые ячейки)", True, 13),
        ("На «01_Настройки»: дни периода продаж, дни периода заказа, коэф. заказа, uplift/downlift, safety stock, пороги.", False, 11),
        ("На «03_Расчёт_заказа»: Остаток, Продажи, Тренд, Коэф. строки, ABC, Цена приходная, Комментарий, Наименование.", False, 11),
        ("Можно добавлять строки ВНИЗУ таблицы, копируя формулы с соседней строки.", False, 11),
        ("", False, 11),
        ("4. Что НЕЛЬЗЯ редактировать (зелёные / расчётные)", True, 13),
        ("Среднедневные, Прогноз, Страховой запас, Потребность, Рекомендуемый заказ, Покрытие, Статус — считаются формулами.", False, 11),
        ("Не удаляйте лист «01_Настройки» и именованную логику ссылок вида Настройки!$B$…", False, 11),
        ("", False, 11),
        ("5. Как изменить период продаж и период заказа", True, 13),
        ("Откройте «01_Настройки» → ячейки B5 (дни анализа продаж) и B6 (дни горизонта заказа).", False, 11),
        ("После изменения Excel пересчитает среднедневные, прогноз и заказ на листе «03_Расчёт_заказа».", False, 11),
        ("", False, 11),
        ("6. Как менять коэффициенты", True, 13),
        ("Глобально: «01_Настройки» — Коэф. заказа, Повышающий, Понижающий, множители ABC.", False, 11),
        ("По строке: колонка «Коэф. строки» на листе расчёта (например 1.2 для акции на конкретный товар).", False, 11),
        ("", False, 11),
        ("7. Светофор и статусы", True, 13),
        ("🔴 Критический дефицит — остаток на исходе, срочный заказ.", False, 11),
        ("🟠 Риск OOS — покрытие меньше порога риска.", False, 11),
        ("🟡 Избыточный запас — слишком большое покрытие.", False, 11),
        ("🟣 Неликвид — нет продаж при наличии остатка.", False, 11),
        ("🔵 Приоритет A — важная позиция к заказу.", False, 11),
        ("🟢 Норма / рост — плановый заказ.", False, 11),
        ("⚪ Запас достаточен — заказ не требуется.", False, 11),
        ("", False, 11),
        ("8. Логика формул (кратко)", True, 13),
        ("Среднедневные = Продажи / ДниПродаж", False, 11),
        ("Прогноз = Среднедневные × ДниЗаказа × Тренд × КоэфЗаказа × Uplift × Downlift × МножABC × КоэфСтроки", False, 11),
        ("Страховой = Среднедневные × ДниSafety(ABC) × Тренд", False, 11),
        ("Потребность = Прогноз + Страховой", False, 11),
        ("Рекомендуемый заказ = MAX(расчётный; MAX(0; МинОстаток - Остаток))", False, 11),
        ("Минимальный остаток по умолчанию — 24 шт на каждый SKU (параметр на «01_Настройки»).", False, 11),
        ("Для неликвидов без продаж расчётный заказ обнуляется, но докупка до минимального остатка сохраняется.", False, 11),
        ("Сумма заказа = Рекомендуемый заказ × Цена приходная.", False, 11),
        ("", False, 11),
        ("9. Советы", True, 13),
        ("Сначала отфильтруйте 🔴 и 🟠, затем класс A, затем остальное.", False, 11),
        ("Длинные наименования переносятся — увеличьте высоту строки при необходимости.", False, 11),
        ("Перед заказом в 1С сверьте единицы измерения и кратность упаковки вручную.", False, 11),
        ]
    )
    ws.column_dimensions["A"].width = 120
    for i, (text, bold, size) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name="Calibri", bold=bold, size=size, color="1F4E79" if bold else "333333")
        cell.alignment = ALIGN_WRAP
        ws.row_dimensions[i].height = 18 if not bold else 22
    ws.freeze_panes = "A2"


def _build_settings(wb: Workbook, meta: Dict[str, Any]) -> None:
    ws = _ws(wb, "01_Настройки")
    ws["A1"] = "НАСТРОЙКИ РАСЧЁТА"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = FILL_HEADER
    ws.merge_cells("A1:C1")

    ws["A2"] = "Жёлтые ячейки — редактируемые. От них зависят формулы на листе «03_Расчёт_заказа»."
    ws.merge_cells("A2:C2")

    rows = [
        (4, "Параметр", "Значение", "Подсказка"),
        (5, "Дни анализа продаж", meta.get("period_days", SETTINGS["default_sales_period_days"]),
         "Делитель для среднедневных продаж"),
        (6, "Дни периода заказа", meta.get("order_period_days", SETTINGS["default_order_period_days"]),
         "Горизонт покрытия / прогноз"),
        (7, "Коэффициент заказа", meta.get("order_coefficient", SETTINGS["order_coefficient"]),
         "Глобальный множитель заказа"),
        (8, "Повышающий коэффициент", meta.get("uplift_coefficient", SETTINGS["uplift_coefficient"]),
         "Акции, сезон, рост трафика"),
        (9, "Понижающий коэффициент", meta.get("downlift_coefficient", SETTINGS["downlift_coefficient"]),
         "Спад, окончание сезона"),
        (10, "Safety stock дни (A)", SETTINGS["safety_stock_days_a"], "Страховой запас для A"),
        (11, "Safety stock дни (B)", SETTINGS["safety_stock_days_b"], "Страховой запас для B"),
        (12, "Safety stock дни (C)", SETTINGS["safety_stock_days_c"], "Страховой запас для C"),
        (13, "Множитель заказа ABC-A", SETTINGS["abc_order_mult_a"], "Приоритет защиты A от OOS"),
        (14, "Множитель заказа ABC-B", SETTINGS["abc_order_mult_b"], "База для B"),
        (15, "Множитель заказа ABC-C", SETTINGS["abc_order_mult_c"], "Осторожный заказ C"),
        (16, "Порог риска OOS, дни", SETTINGS["oos_risk_cover_days"], "Покрытие ниже = риск"),
        (17, "Критический OOS, дни", SETTINGS["critical_oos_cover_days"], "Покрытие ниже = критично"),
        (18, "Порог избытка, дни", SETTINGS["overstock_cover_days"], "Покрытие выше = избыток"),
        (19, "Дни без продаж → неликвид", SETTINGS["dead_stock_days_no_sales"], "Критерий неликвида"),
        (20, "Минимальный остаток, шт", SETTINGS.get("min_stock_target", 24),
         "Докупка заказа до этого уровня по каждому SKU (даже при нулевом остатке / низких продажах)"),
        (21, "Дата начала периода", str(meta.get("date_from", ""))[:10], "Информативно"),
        (22, "Дата окончания периода", str(meta.get("date_to", ""))[:10], "Информативно"),
    ]

    for r, a, b, c in rows:
        ws.cell(row=r, column=1, value=a).border = THIN
        cell_b = ws.cell(row=r, column=2, value=b)
        cell_b.border = THIN
        ws.cell(row=r, column=3, value=c).border = THIN
        if r == 4:
            for col in range(1, 4):
                ws.cell(row=r, column=col).fill = FILL_HEADER
                ws.cell(row=r, column=col).font = FONT_HEADER
        elif r <= 20:
            cell_b.fill = FILL_EDIT
            cell_b.comment = Comment(c, "Система", width=200, height=50)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45

    ws["A24"] = "Справка по формулам"
    ws["A24"].font = Font(bold=True, size=12, color="1F4E79")
    ws["A25"] = "Прогноз = (Продажи/B5)*B6*Тренд*B7*B8*B9*МножABC*КоэфСтроки"
    ws["A26"] = "Заказ = MAX(расчётный; MAX(0; B20 - Остаток)) - докупка до минимального остатка"


def _build_dashboard(wb: Workbook, df: pd.DataFrame, meta: Dict[str, Any]) -> None:
    ws = _ws(wb, "02_Дашборд")
    ws["A1"] = "ИТОГОВАЯ АНАЛИТИКА ЗАКУПОК"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = FILL_HEADER
    ws.merge_cells("A1:D1")

    cards = [
        (3, "Позиций в расчёте", meta.get("items_count", len(df))),
        (4, "Строк к заказу", meta.get("order_lines", 0)),
        (5, "Суммарный рекомендуемый заказ, ед.", round(meta.get("order_qty_total", 0), 2)),
        (6, "Рисков OOS", meta.get("oos_count", 0)),
        (7, "Неликвидов", meta.get("dead_count", 0)),
        (8, "Избыточных запасов", meta.get("overstock_count", 0)),
        (9, "Класс A / B / C", f"{meta.get('abc_a_count', 0)} / {meta.get('abc_b_count', 0)} / {meta.get('abc_c_count', 0)}"),
        (10, "Период продаж, дни", meta.get("period_days", "")),
        (11, "Горизонт заказа, дни", meta.get("order_period_days", "")),
        (12, "Продажи за период, ед.", round(meta.get("total_sales_qty", 0), 2)),
        (13, "Текущий остаток, ед.", round(meta.get("total_stock", 0), 2)),
        (14, "Детализация", "по магазинам" if meta.get("grain") == GRAIN_STORE else "сводно по сети"),
        (15, "Магазинов в расчёте", meta.get("store_count", 0)),
        (16, "Сумма заказа, ₽", round(meta.get("order_sum_total", 0), 2)),
    ]
    ws["A2"] = "Показатель"
    ws["B2"] = "Значение"
    style_header_row(ws, 2, 2)
    for r, title, val in cards:
        ws.cell(row=r, column=1, value=title).border = THIN
        cell = ws.cell(row=r, column=2, value=val)
        cell.border = THIN
        cell.fill = FILL_FORMULA

    # Топ приоритетов
    ws["A15"] = "ТОП приоритетных позиций (критично / OOS / A)"
    ws["A15"].font = Font(bold=True, size=12, color="1F4E79")
    top = df.sort_values("priority").head(15)
    headers = ["Артикул", "Наименование", "ABC", "Остаток", "Заказ", "Статус", "Светофор"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=16, column=i, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
    for r_i, (_, row) in enumerate(top.iterrows(), start=17):
        vals = [
            row["sku"],
            row["name"],
            row["abc_class"],
            row["stock"],
            row["recommended_order"],
            row["status"],
            row["traffic_light"],
        ]
        for c_i, v in enumerate(vals, 1):
            cell = ws.cell(row=r_i, column=c_i, value=v if not isinstance(v, float) else round(v, 3))
            cell.border = THIN
            if c_i == 2:
                cell.alignment = ALIGN_WRAP

    # ABC counts for chart
    ws["F3"] = "ABC"
    ws["G3"] = "Кол-во"
    ws["F4"] = "A"
    ws["G4"] = meta.get("abc_a_count", 0)
    ws["F5"] = "B"
    ws["G5"] = meta.get("abc_b_count", 0)
    ws["F6"] = "C"
    ws["G6"] = meta.get("abc_c_count", 0)
    chart = BarChart()
    chart.title = "Распределение ABC"
    chart.dataLabels = None
    data = Reference(ws, min_col=7, min_row=3, max_row=6)
    cats = Reference(ws, min_col=6, min_row=4, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 12
    chart.height = 8
    ws.add_chart(chart, "F8")

    autosize_columns(ws, name_col=2)
    ws.freeze_panes = "A3"


def _abc_mult_formula(row: int) -> str:
    abc = _col_letter("abc")
    return (
        f'IF({abc}{row}="A",\'01_Настройки\'!$B$13,'
        f'IF({abc}{row}="B",\'01_Настройки\'!$B$14,\'01_Настройки\'!$B$15))'
    )


def _safety_days_formula(row: int) -> str:
    abc = _col_letter("abc")
    return (
        f'IF({abc}{row}="A",\'01_Настройки\'!$B$10,'
        f'IF({abc}{row}="B",\'01_Настройки\'!$B$11,\'01_Настройки\'!$B$12))'
    )


def _status_formula(row: int) -> str:
    sales = _col_letter("sales")
    cover = _col_letter("cover")
    stock = _col_letter("stock")
    rec = _col_letter("rec_order")
    abc = _col_letter("abc")
    trend = _col_letter("trend")
    return (
        f'IF(OR(AND({sales}{row}>0,{cover}{row}<\'01_Настройки\'!$B$17),AND({sales}{row}>0,{stock}{row}<=0)),"Критический дефицит",'
        f'IF(AND({sales}{row}>0,{cover}{row}<\'01_Настройки\'!$B$16),"Риск out-of-stock",'
        f'IF(AND({sales}{row}<=0,{stock}{row}>0),"Неликвид / зависший",'
        f'IF(AND({sales}{row}>0,{cover}{row}>\'01_Настройки\'!$B$18),"Избыточный запас",'
        f'IF(AND({rec}{row}>0,{abc}{row}="A"),"Приоритетный заказ (A)",'
        f'IF(AND({rec}{row}>0,{trend}{row}>=1.15),"Рост спроса — усилить заказ",'
        f'IF({rec}{row}>0,"К заказу","Запас достаточен")))))))'
    )


def _risk_formula(row: int) -> str:
    st = _col_letter("status")
    return (
        f'IF({st}{row}="Критический дефицит","Высокий риск OOS",'
        f'IF({st}{row}="Риск out-of-stock","Средний риск OOS",'
        f'IF({st}{row}="Неликвид / зависший","Замороженный капитал",'
        f'IF({st}{row}="Избыточный запас","Завышенный stock",'
        f'IF({st}{row}="Приоритетный заказ (A)","Контроль наличия",'
        f'IF({st}{row}="Рост спроса — усилить заказ","Тренд вверх",'
        f'IF({st}{row}="К заказу","Норма","Низкий")))))))'
    )


def _light_formula(row: int) -> str:
    st = _col_letter("status")
    return (
        f'IF({st}{row}="Критический дефицит","🔴",'
        f'IF({st}{row}="Риск out-of-stock","🟠",'
        f'IF({st}{row}="Неликвид / зависший","🟣",'
        f'IF({st}{row}="Избыточный запас","🟡",'
        f'IF({st}{row}="Приоритетный заказ (A)","🔵",'
        f'IF(OR({st}{row}="К заказу",{st}{row}="Рост спроса — усилить заказ"),"🟢","⚪"))))))'
    )


def _build_main_calc(wb: Workbook, df: pd.DataFrame, meta: Dict[str, Any]) -> None:
    ws = _ws(wb, "03_Расчёт_заказа")
    grain_label = "по магазинам" if meta.get("grain") == GRAIN_STORE else "сводно по сети"
    ws["A1"] = (
        f"ОСНОВНОЙ РАСЧЁТ ЗАКАЗА  |  {grain_label}  |  Жёлтые — ввод  |  Зелёные — формулы  |  "
        "Параметры: лист 01_Настройки"
    )
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = FILL_HEADER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN

    tips = {
        COL["stock"]: "Редактируемый остаток",
        COL["sales"]: "Редактируемые продажи за период",
        COL["trend"]: "Тренд: >1 рост, <1 спад",
        COL["line_coef"]: "Локальный множитель строки",
        COL["abc"]: "Класс ABC (можно скорректировать)",
        COL["quantum"]: "Квант упаковки (шт). Заказ и перемещения округляются вверх до кванта.",
        COL["rec_order"]: "Итоговый заказ (формула + округление до кванта в модуле)",
        COL["price"]: "Цена приходная из справочника, можно поправить",
        COL["note"]: "Свободный комментарий",
    }
    for col, tip in tips.items():
        ws.cell(row=2, column=col).comment = Comment(tip, "Система", width=180, height=40)

    sales_l = _col_letter("sales")
    stock_l = _col_letter("stock")
    trend_l = _col_letter("trend")
    coef_l = _col_letter("line_coef")
    avg_l = _col_letter("avg_daily")
    fc_l = _col_letter("forecast")
    sf_l = _col_letter("safety")
    need_l = _col_letter("need")
    raw_l = _col_letter("raw_order")
    rec_l = _col_letter("rec_order")
    price_l = _col_letter("price")

    n = len(df)
    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 3
        ws.cell(row=r, column=COL["id"], value=int(row["row_id"]))
        ws.cell(row=r, column=COL["store"], value=safe_str(row.get("store", NETWORK_STORE_LABEL)))
        ws.cell(row=r, column=COL["supplier"], value=safe_str(row.get("supplier_name", "")))
        ws.cell(row=r, column=COL["sku"], value=safe_str(row["sku"]))
        name_cell = ws.cell(row=r, column=COL["name"], value=safe_str(row["name"]))
        name_cell.alignment = ALIGN_WRAP
        ws.cell(row=r, column=COL["barcode"], value=safe_str(row.get("barcode", "")))
        ws.cell(row=r, column=COL["uom"], value=safe_str(row.get("uom", "")))
        q_cell = ws.cell(row=r, column=COL["quantum"], value=int(row.get("quantum", 1) or 1))
        q_cell.fill = FILL_EDIT

        stock_cell = ws.cell(row=r, column=COL["stock"], value=float(row["stock"]))
        sales_cell = ws.cell(row=r, column=COL["sales"], value=float(row["sales_qty"]))
        trend_cell = ws.cell(row=r, column=COL["trend"], value=round(float(row["trend_coef"]), 4))
        coef_cell = ws.cell(row=r, column=COL["line_coef"], value=1.0)
        abc_cell = ws.cell(row=r, column=COL["abc"], value=str(row["abc_class"]))
        price_cell = ws.cell(row=r, column=COL["price"], value=float(row.get("purchase_price", 0) or 0))
        price_cell.number_format = "#,##0.00"

        for cell in (stock_cell, sales_cell, trend_cell, coef_cell, abc_cell, name_cell, price_cell, q_cell):
            cell.fill = FILL_EDIT
            cell.border = THIN

        # Заказ: базовая формула, затем округление вверх до кванта (столбец quantum).
        q_l = _col_letter("quantum")
        base_order = (
            f"MAX(IF(AND({sales_l}{r}<=0,{trend_l}{r}<=1),0,MAX(0,CEILING({raw_l}{r},1))),"
            f"MAX(0,CEILING('01_Настройки'!$B$20-{stock_l}{r},1)))"
        )
        ws.cell(row=r, column=COL["avg_daily"], value=f"=IF('01_Настройки'!$B$5=0,0,{sales_l}{r}/'01_Настройки'!$B$5)")
        ws.cell(
            row=r,
            column=COL["forecast"],
            value=(
                f"={avg_l}{r}*'01_Настройки'!$B$6*{trend_l}{r}*'01_Настройки'!$B$7*"
                f"'01_Настройки'!$B$8*'01_Настройки'!$B$9*{_abc_mult_formula(r)}*{coef_l}{r}"
            ),
        )
        ws.cell(row=r, column=COL["safety"], value=f"={avg_l}{r}*{_safety_days_formula(r)}*{trend_l}{r}")
        ws.cell(row=r, column=COL["need"], value=f"={fc_l}{r}+{sf_l}{r}")
        ws.cell(row=r, column=COL["raw_order"], value=f"={need_l}{r}-{stock_l}{r}")
        store_name = safe_str(row.get("store", ""))
        if is_central_warehouse(store_name) and float(row.get("supplier_order_qty", 0) or 0) > 0:
            # Заказ на ЦС = сумма потребности магазинов (считается в модуле), не локальная формула.
            rec_cell = ws.cell(
                row=r,
                column=COL["rec_order"],
                value=float(row.get("recommended_order", 0) or 0),
            )
            rec_cell.fill = FILL_EDIT
        else:
            ws.cell(
                row=r,
                column=COL["rec_order"],
                value=(
                    f'=IF(({base_order})<=0,0,'
                    f'IF(OR({q_l}{r}="",{q_l}{r}<=1),{base_order},'
                    f'CEILING({base_order},{q_l}{r})))'
                ),
            )
        ws.cell(
            row=r,
            column=COL["cover"],
            value=f"=IF({avg_l}{r}>0,{stock_l}{r}/{avg_l}{r},IF({stock_l}{r}>0,9999,0))",
        )
        ws.cell(row=r, column=COL["status"], value=f"={_status_formula(r)}")
        ws.cell(row=r, column=COL["risk"], value=f"={_risk_formula(r)}")
        ws.cell(row=r, column=COL["light"], value=f"={_light_formula(r)}")
        sum_cell = ws.cell(row=r, column=COL["order_sum"], value=f"={rec_l}{r}*{price_l}{r}")
        sum_cell.number_format = "#,##0.00"
        note = ws.cell(row=r, column=COL["note"], value="")
        tin = float(row.get("transfer_in", 0) or 0)
        tout = float(row.get("transfer_out", 0) or 0)
        if tin > 0:
            note.value = f"Переместить на точку: +{tin:g}"
        elif tout > 0:
            note.value = f"Списать со склада: −{tout:g}"
        note.fill = FILL_EDIT

        formula_cols = {
            COL["avg_daily"], COL["forecast"], COL["safety"], COL["need"], COL["raw_order"],
            COL["rec_order"], COL["cover"], COL["status"], COL["risk"], COL["light"], COL["order_sum"],
        }
        wrap_cols = {COL["name"], COL["store"], COL["supplier"]}
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN
            if c in formula_cols:
                cell.fill = FILL_FORMULA
            cell.alignment = ALIGN_WRAP if c in wrap_cols else ALIGN_CENTER

        name_len = len(safe_str(row["name"]))
        ws.row_dimensions[r].height = max(18, min(60, 14 + name_len // 40 * 12))

    last_row = n + 2
    abc_l = _col_letter("abc")
    status_l = _col_letter("status")
    rec_l = _col_letter("rec_order")
    if n > 0:
        table_ref = f"A2:{get_column_letter(len(HEADERS))}{last_row}"
        table = Table(displayName="РасчётЗаказа", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name=SETTINGS["table_style"],
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        add_abc_validation(ws, f"{abc_l}3:{abc_l}{last_row}")
        add_oos_conditional(ws, status_l, 3, last_row)
        add_order_highlight(ws, rec_l, 3, last_row)
    else:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{max(last_row, 2)}"

    ws.freeze_panes = "F3"
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        if letter in {"B", "C", "E"}:
            continue
        ws.column_dimensions[letter].width = 14 if col > 3 else 12
    ws.column_dimensions[_col_letter("status")].width = 28
    ws.column_dimensions[_col_letter("risk")].width = 22
    ws.column_dimensions[_col_letter("note")].width = 24
    ws.column_dimensions[_col_letter("barcode")].width = 16


def _build_abc_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = _ws(wb, "04_ABC")
    ws["A1"] = "ABC-АНАЛИЗ (по доле продаж)"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = FILL_HEADER
    ws.merge_cells("A1:H1")

    cols = [
        "sku",
        "name",
        "abc_class",
        "abc_rank",
        "sales_qty",
        "sales_amount",
        "abc_share",
        "abc_cum_share",
    ]
    titles = [
        "Артикул",
        "Наименование",
        "ABC",
        "Ранг",
        "Продажи, ед.",
        "Продажи, сумма",
        "Доля",
        "Накопленная доля",
    ]
    for c, t in enumerate(titles, 1):
        cell = ws.cell(row=2, column=c, value=t)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN

    view = df.sort_values("abc_rank") if "abc_rank" in df.columns else df
    for i, (_, row) in enumerate(view.iterrows(), start=3):
        values = [
            row["sku"],
            row["name"],
            row["abc_class"],
            int(row.get("abc_rank", i - 2)),
            float(row["sales_qty"]),
            float(row.get("sales_amount", 0) or 0),
            round(float(row.get("abc_share", 0) or 0), 6),
            round(float(row.get("abc_cum_share", 0) or 0), 6),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = THIN
            if c == 2:
                cell.alignment = ALIGN_WRAP
            if c == 3:
                abc = str(v)
                if abc == "A":
                    cell.fill = PatternFill("solid", fgColor=SETTINGS["color_a"])
                elif abc == "B":
                    cell.fill = PatternFill("solid", fgColor=SETTINGS["color_b"])
                else:
                    cell.fill = PatternFill("solid", fgColor=SETTINGS["color_c"])

    last = 2 + len(view)
    if len(view):
        ws.auto_filter.ref = f"A2:H{last}"
    ws.freeze_panes = "A3"
    autosize_columns(ws, name_col=2)


def _build_filtered_sheet(wb: Workbook, subset: pd.DataFrame, title: str, header: str) -> None:
    ws = _ws(wb, title)
    ws["A1"] = header
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = FILL_HEADER
    ws.merge_cells("A1:J1")

    titles = [
        "Артикул",
        "Наименование",
        "ABC",
        "Остаток",
        "Продажи",
        "Покрытие, дни",
        "Рек. заказ",
        "Статус",
        "Риск",
        "Светофор",
    ]
    for c, t in enumerate(titles, 1):
        cell = ws.cell(row=2, column=c, value=t)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN

    subset = subset.sort_values("priority") if "priority" in subset.columns else subset
    for i, (_, row) in enumerate(subset.iterrows(), start=3):
        vals = [
            row.get("sku", ""),
            row.get("name", ""),
            row.get("abc_class", ""),
            float(row.get("stock", 0) or 0),
            float(row.get("sales_qty", 0) or 0),
            round(float(row.get("cover_days", 0) or 0), 2),
            float(row.get("recommended_order", 0) or 0),
            row.get("status", ""),
            row.get("risk_level", ""),
            row.get("traffic_light", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = THIN
            if c == 2:
                cell.alignment = ALIGN_WRAP

    last = max(2 + len(subset), 2)
    ws.auto_filter.ref = f"A2:J{last}"
    ws.freeze_panes = "A3"
    if subset.empty:
        ws["A3"] = "Нет позиций в этой категории — отличный знак."
    autosize_columns(ws, name_col=2)


def _build_trends_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = _ws(wb, "08_Тренды")
    ws["A1"] = "ТРЕНДЫ ПРОДАЖ (1-я vs 2-я половина периода)"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = FILL_HEADER
    ws.merge_cells("A1:I1")

    titles = [
        "Артикул",
        "Наименование",
        "ABC",
        "Продажи 1 пол.",
        "Продажи 2 пол.",
        "Тренд",
        "Метка",
        "Рек. заказ",
        "Статус",
    ]
    for c, t in enumerate(titles, 1):
        cell = ws.cell(row=2, column=c, value=t)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN

    view = df.sort_values("trend_coef", ascending=False)
    for i, (_, row) in enumerate(view.iterrows(), start=3):
        vals = [
            row["sku"],
            row["name"],
            row["abc_class"],
            float(row.get("sales_h1", 0) or 0),
            float(row.get("sales_h2", 0) or 0),
            round(float(row.get("trend_coef", 1) or 1), 4),
            row.get("trend_label", ""),
            float(row.get("recommended_order", 0) or 0),
            row.get("status", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = THIN
            if c == 2:
                cell.alignment = ALIGN_WRAP
            if c == 7:
                label = str(v)
                if label == "Рост":
                    cell.fill = PatternFill("solid", fgColor=SETTINGS["color_ok"])
                elif label == "Спад":
                    cell.fill = PatternFill("solid", fgColor=SETTINGS["color_warning"])

    last = max(2 + len(view), 2)
    ws.auto_filter.ref = f"A2:I{last}"
    ws.freeze_panes = "A3"
    autosize_columns(ws, name_col=2)


def _build_supplier_order_sheet(
    wb: Workbook,
    subset: pd.DataFrame,
    supplier_name: str | None,
    grain: str = "network",
) -> None:
    """Упрощённый заказ: наименование, количество, штрихкод, цена, сумма. Всегда."""
    ws = _ws(wb, "09_Заказ_поставщику")
    title = "ЗАКАЗ ПОСТАВЩИКУ"
    if supplier_name:
        title += f": {supplier_name}"
    elif grain == GRAIN_STORE:
        title += " (все контрагенты, по магазинам)"
    else:
        title += " (все контрагенты, сводно)"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = FILL_HEADER

    include_store = grain == GRAIN_STORE
    include_supplier = not bool(supplier_name)
    titles = ["№"]
    if include_store:
        titles.append("Магазин")
    if include_supplier:
        titles.append("Поставщик")
    titles.extend(["Наименование", "Штрихкод", "Ед.", "Квант", "Заказ, кол-во", "Цена", "Сумма"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(titles))

    ws["A2"] = (
        "Заказ округлён до кванта упаковки. "
        + (
            "В режиме магазинов — строки по точкам с заказом > 0. "
            "Адлер/Сочи и др. с заказом 0 смотрите на 03/10 (запас достаточен). "
            "Сумма по магазинам = заказ на пополнение «Склад основной1». "
            if grain == GRAIN_STORE
            else ""
        )
        + "Количество можно править — сумму пересчитайте как кол-во × цена."
    )
    ws["A2"].alignment = ALIGN_WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(titles))

    for c, t in enumerate(titles, 1):
        cell = ws.cell(row=3, column=c, value=t)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN

    sort_by = [c for c in (("store",) if include_store else ()) + (("supplier_name",) if include_supplier else ()) + ("name",)]
    view = subset.copy()
    if not view.empty:
        existing = [c for c in sort_by if c in view.columns]
        if existing:
            view = view.sort_values(existing)

    total_qty = 0.0
    total_sum = 0.0
    n = 0
    for _, row in view.iterrows():
        qty = float(row.get("recommended_order", 0) or 0)
        if qty <= 0:
            continue
        n += 1
        price = float(row.get("purchase_price", 0) or 0)
        amount = round(qty * price, 2)
        total_qty += qty
        total_sum += amount
        vals: list = [n]
        if include_store:
            vals.append(safe_str(row.get("store", "")))
        if include_supplier:
            vals.append(safe_str(row.get("supplier_name", "")))
        vals.extend(
            [
                safe_str(row.get("name", "")),
                safe_str(row.get("barcode", "")),
                safe_str(row.get("uom", "")),
                int(row.get("quantum", 1) or 1),
                qty,
                price,
                amount,
            ]
        )
        r = n + 3
        name_idx = 1 + int(include_store) + int(include_supplier) + 1
        # ... Квант, Заказ, Цена, Сумма
        quantum_idx = len(titles) - 3
        qty_idx = len(titles) - 2
        price_idx = len(titles) - 1
        sum_idx = len(titles)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            if c == name_idx:
                cell.alignment = ALIGN_WRAP
            if c == quantum_idx:
                cell.fill = FILL_EDIT
            if c == qty_idx:
                cell.fill = FILL_EDIT
            if c in (price_idx, sum_idx):
                cell.fill = FILL_FORMULA
                cell.number_format = "#,##0.00"
        ws.row_dimensions[r].height = 28

    last = 3 + max(n, 1)
    tot = last + 1 if n else 4
    if n:
        tot = n + 4
        ws.cell(row=tot, column=1, value="ИТОГО")
        for c in range(1, len(titles) + 1):
            ws.cell(row=tot, column=c).fill = FILL_HEADER
            ws.cell(row=tot, column=c).font = FONT_HEADER
            ws.cell(row=tot, column=c).border = THIN
        ws.cell(row=tot, column=len(titles) - 2, value=total_qty)
        sum_cell = ws.cell(row=tot, column=len(titles), value=round(total_sum, 2))
        sum_cell.number_format = "#,##0.00"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(titles))}{n + 3}"
        ws.freeze_panes = "A4"
    else:
        ws["A4"] = "Нет позиций с рекомендуемым заказом > 0."
        ws.auto_filter.ref = f"A3:{get_column_letter(len(titles))}4"
        ws.freeze_panes = "A4"

    name_col = 2 + int(include_store) + int(include_supplier)
    autosize_columns(ws, name_col=name_col)
    ws.column_dimensions["A"].width = 6


def _build_store_matrix(wb: Workbook, subset: pd.DataFrame) -> None:
    """Шахматка заказ SKU × магазин."""
    ws = _ws(wb, "10_Матрица_заказ")
    if subset is None or subset.empty or "store" not in subset.columns:
        ws["A1"] = "Нет данных для матрицы по магазинам."
        return
    stores = sorted(
        {
            safe_str(x)
            for x in subset["store"].tolist()
            if safe_str(x) and not is_central_warehouse(x)
        }
    )
    headers = ["Поставщик", "Наименование", "Штрихкод", "Квант"] + stores + ["Итого заказ", "Сумма, ₽"]
    ws["A1"] = "ЗАКАЗ SKU × МАГАЗИН (кванты)"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = FILL_HEADER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A2"] = "Количество в точках уже кратно кванту. Пустая ячейка = заказ 0. Центральный склад в матрицу не входит."
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    for c, t in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=t)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
        cell.alignment = ALIGN_CENTER

    grouped = {}
    for _, row in subset.iterrows():
        key = safe_str(row.get("sku_key") or row.get("sku") or row.get("name"))
        grouped.setdefault(
            key,
            {
                "supplier": safe_str(row.get("supplier_name", "")),
                "name": safe_str(row.get("name", "")),
                "barcode": safe_str(row.get("barcode", "")),
                "quantum": int(row.get("quantum", 1) or 1),
                "price": float(row.get("purchase_price", 0) or 0),
                "qty": {s: 0.0 for s in stores},
            },
        )
        store = safe_str(row.get("store", ""))
        qty = float(row.get("recommended_order", 0) or 0)
        if store in grouped[key]["qty"]:
            grouped[key]["qty"][store] += qty

    r = 4
    for item in sorted(grouped.values(), key=lambda x: (x["supplier"], x["name"])):
        total = sum(item["qty"].values())
        vals = [item["supplier"], item["name"], item["barcode"], item["quantum"]]
        vals.extend(item["qty"][s] or None for s in stores)
        vals.extend([total, round(total * item["price"], 2)])
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            if c in (1, 2):
                cell.alignment = ALIGN_WRAP
            if c == 4:
                cell.fill = FILL_EDIT
            if 5 <= c <= 4 + len(stores) and v:
                cell.fill = FILL_EDIT
            if c >= len(vals) - 1:
                cell.fill = FILL_FORMULA
            if c == len(vals):
                cell.number_format = "#,##0.00"
        ws.row_dimensions[r].height = 28
        r += 1
    ws.freeze_panes = "E4"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10


def _build_transfer_sheet(
    wb: Workbook,
    transfers: pd.DataFrame | None,
    meta: Dict[str, Any],
) -> None:
    """Лист перемещений: Склад основной → магазины (до заказа поставщику)."""
    ws = _ws(wb, "11_Перемещение_со_склада")
    ws["A1"] = "ПЕРЕМЕСТИТЬ СО СКЛАДА В МАГАЗИН (до заказа поставщику)"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = FILL_HEADER

    titles = [
        "№",
        "Со склада",
        "В магазин",
        "Приоритет",
        "Артикул",
        "Наименование",
        "Штрихкод",
        "Ед.",
        "Квант",
        "Кол-во переместить",
        "Потребность до",
        "Остаток в магазине до",
        "Остаток на складе до",
        "Остаток на складе после",
        "Заказ поставщику после",
        "Продажи SKU в точке",
        "Поставщик",
        "Цена",
        "Сумма (справка)",
    ]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(titles))
    ws["A2"] = (
        "Перемещение только полными квантами. "
        "Сначала закрываем потребность магазинов остатком «Склад основной1». "
        "Приоритет: Флагман, далее точки по убыванию продаж SKU. "
        "Лист 09_Заказ_поставщику = сумма остаточной потребности магазинов на ЦС. "
        f"Строк перемещения: {int(meta.get('transfer_lines', 0) or 0)}, "
        f"всего шт: {float(meta.get('transfer_qty_total', 0) or 0):.0f}."
    )
    ws["A2"].alignment = ALIGN_WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(titles))
    ws.row_dimensions[2].height = 36

    for c, t in enumerate(titles, 1):
        cell = ws.cell(row=3, column=c, value=t)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
        cell.alignment = ALIGN_CENTER

    if transfers is None or transfers.empty:
        ws["A4"] = (
            "Нет предложений по перемещению: на центральном складе нет остатка "
            "под текущую потребность магазинов, либо детализация не по магазинам."
        )
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(titles))
        autosize_columns(ws, name_col=6)
        return

    view = transfers.sort_values(
        ["priority_rank", "to_store", "name"],
        ascending=[True, True, True],
    )
    total_qty = 0.0
    total_sum = 0.0
    n = 0
    for _, row in view.iterrows():
        qty = float(row.get("transfer_qty", 0) or 0)
        if qty <= 0:
            continue
        n += 1
        price = float(row.get("purchase_price", 0) or 0)
        amount = round(qty * price, 2)
        total_qty += qty
        total_sum += amount
        vals = [
            n,
            safe_str(row.get("from_store", "Склад основной1")),
            safe_str(row.get("to_store", "")),
            int(row.get("priority_rank", 0) or 0),
            safe_str(row.get("sku", "")),
            safe_str(row.get("name", "")),
            safe_str(row.get("barcode", "")),
            safe_str(row.get("uom", "")),
            int(row.get("quantum", 1) or 1),
            qty,
            float(row.get("need_before", 0) or 0),
            float(row.get("store_stock_before", 0) or 0),
            float(row.get("central_stock_before", 0) or 0),
            float(row.get("central_stock_after", 0) or 0),
            float(row.get("order_after", 0) or 0),
            float(row.get("sales_qty", 0) or 0),
            safe_str(row.get("supplier_name", "")),
            price,
            amount,
        ]
        r = n + 3
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            if c == 6:
                cell.alignment = ALIGN_WRAP
            if c in (9, 10):
                cell.fill = FILL_EDIT
            if c in (18, 19):
                cell.fill = FILL_FORMULA
                cell.number_format = "#,##0.00"
        ws.row_dimensions[r].height = 28

    tot = n + 4
    ws.cell(row=tot, column=1, value="ИТОГО")
    for c in range(1, len(titles) + 1):
        ws.cell(row=tot, column=c).fill = FILL_HEADER
        ws.cell(row=tot, column=c).font = FONT_HEADER
        ws.cell(row=tot, column=c).border = THIN
    ws.cell(row=tot, column=10, value=total_qty)
    sum_cell = ws.cell(row=tot, column=19, value=round(total_sum, 2))
    sum_cell.number_format = "#,##0.00"

    ws.auto_filter.ref = f"A3:{get_column_letter(len(titles))}{n + 3}"
    ws.freeze_panes = "A4"
    autosize_columns(ws, name_col=6)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["F"].width = 55
