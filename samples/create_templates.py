"""
Создаёт Excel-шаблоны для загрузки остатков и продаж.

Структура каждого файла:
  лист «Данные»      — таблица для заполнения (именно его читает модуль);
  лист «Инструкция»  — памятка пользователю (модуль этот лист игнорирует).

Запуск:
    python samples/create_templates.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SAMPLES = Path(__file__).resolve().parent
TEMPLATES = SAMPLES / "templates"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="FFF2CC")
HINT_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="center")


def _style_header(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_instruction(ws, lines: list[tuple[str, bool, int]]) -> None:
    ws.column_dimensions["A"].width = 110
    for i, (text, bold, size) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(
            name="Calibri",
            bold=bold,
            size=size,
            color="1F4E79" if bold else "333333",
        )
        cell.alignment = WRAP
        ws.row_dimensions[i].height = 18 if size <= 11 else 22
    ws.freeze_panes = "A2"


def create_stock_template(path: Path) -> Path:
    wb = Workbook()

    # --- Данные (первый лист — на случай старых версий загрузчика) ---
    ws = wb.active
    ws.title = "Данные"
    headers = ["Артикул", "Наименование", "Остаток", "Ед. изм.", "Склад"]
    _style_header(ws, headers)

    tips = {
        1: "Код/артикул товара из 1С. Лучше как текст, чтобы не терялись нули.",
        2: "Полное наименование. Можно длинное — не обрезается.",
        3: "Количество на складе (число). Можно с дробями.",
        4: "Необязательно: шт, кг, л и т.д.",
        5: "Необязательно: название склада.",
    }
    for col, tip in tips.items():
        ws.cell(row=1, column=col).comment = Comment(tip, "Шаблон", width=220, height=60)

    # Примеры (жёлтые) — можно удалить или заменить своими данными
    examples = [
        ["A-001", "Молоко ультрапастеризованное 3,2% 1л", 12, "л", "Основной"],
        ["A-002", "Хлеб пшеничный нарезка 500г", 40, "шт", "Основной"],
        ["A-003", "Куриное филе охлаждённое 1кг", 5, "кг", "Холодильник"],
    ]
    for r, row in enumerate(examples, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN
            if c == 2:
                cell.alignment = WRAP
        ws.row_dimensions[r].height = 30

    # Пустые строки для ввода
    for r in range(5, 55):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c, value=None)
            cell.border = THIN
            if c == 3:
                cell.number_format = "0.###"

    widths = [14, 55, 12, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Подсказку текстом в колонку «Артикул» не пишем — иначе попадёт в загрузку.
    # Вся инструкция только на листе «Инструкция».

    # --- Инструкция ---
    wi = wb.create_sheet("Инструкция")
    _write_instruction(
        wi,
        [
            ("ШАБЛОН ОСТАТКОВ — ИНСТРУКЦИЯ", True, 16),
            ("", False, 11),
            ("1. Куда вносить данные", True, 13),
            (
                "Только на лист «Данные». Первая строка — заголовки, со второй строки — товары. "
                "Этот лист «Инструкция» модуль при загрузке полностью игнорирует.",
                False,
                11,
            ),
            ("", False, 11),
            ("2. Обязательные колонки", True, 13),
            ("Артикул — код товара (как в 1С).", False, 11),
            ("Наименование — полное название товара.", False, 11),
            ("Остаток — количество на складе (число).", False, 11),
            ("", False, 11),
            ("3. Необязательные колонки", True, 13),
            ("Ед. изм. — единица измерения.", False, 11),
            ("Склад — склад хранения.", False, 11),
            ("", False, 11),
            ("4. Как заполнять", True, 13),
            ("Скопируйте данные из выгрузки 1С в лист «Данные» под правильные заголовки.", False, 11),
            ("Или выгрузите отчёт 1С и перенесите столбцы так, чтобы названия совпали со строкой 1.", False, 11),
            ("Длинные наименования оставляйте целиком — не сокращайте.", False, 11),
            ("Артикул лучше вводить как текст (если начинается с нуля).", False, 11),
            ("Пустые строки внизу можно не удалять — модуль их пропустит.", False, 11),
            ("", False, 11),
            ("5. Чего нельзя делать", True, 13),
            ("Не переименовывайте лист «Данные».", False, 11),
            ("Не меняйте названия заголовков в первой строке (или добавьте синоним в column_mapping.py).", False, 11),
            ("Не вставляйте текст инструкции на лист «Данные» над таблицей.", False, 11),
            ("Не объединяйте ячейки в области таблицы данных.", False, 11),
            ("", False, 11),
            ("6. Как загрузить в модуль", True, 13),
            ("Сохраните файл (Excel → Сохранить).", False, 11),
            ("Запустите программу → шаг «Загрузите файл остатков» → выберите этот файл.", False, 11),
            ("", False, 11),
            ("7. Примеры в файле", True, 13),
            ("Жёлтые строки на листе «Данные» — учебные примеры. Перед реальной работой замените или удалите их.", False, 11),
        ],
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def create_sales_template(path: Path) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "Данные"
    headers = ["Артикул", "Наименование", "Дата", "Количество", "Сумма"]
    _style_header(ws, headers)

    tips = {
        1: "Код/артикул товара.",
        2: "Полное наименование.",
        3: "Дата продажи. Формат ДД.ММ.ГГГГ или ГГГГ-ММ-ДД.",
        4: "Продано за день / документ (число).",
        5: "Необязательно: сумма продажи в рублях.",
    }
    for col, tip in tips.items():
        ws.cell(row=1, column=col).comment = Comment(tip, "Шаблон", width=220, height=50)

    examples = [
        ["A-001", "Молоко ультрапастеризованное 3,2% 1л", "01.06.2026", 2, 200],
        ["A-001", "Молоко ультрапастеризованное 3,2% 1л", "02.06.2026", 3, 300],
        ["A-002", "Хлеб пшеничный нарезка 500г", "01.06.2026", 5, 250],
        ["A-003", "Куриное филе охлаждённое 1кг", "03.06.2026", 1.5, 450],
    ]
    for r, row in enumerate(examples, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN
            if c == 2:
                cell.alignment = WRAP
            if c == 3:
                cell.number_format = "DD.MM.YYYY"
            if c in (4, 5):
                cell.number_format = "0.###"
        ws.row_dimensions[r].height = 30

    for r in range(6, 106):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c, value=None)
            cell.border = THIN
            if c == 3:
                cell.number_format = "DD.MM.YYYY"
            if c in (4, 5):
                cell.number_format = "0.###"

    widths = [14, 55, 14, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Инструкция только на отдельном листе — не на «Данные».

    # Дата: подсказка формата через validation не обязательна, но добавим комментарий
    wi = wb.create_sheet("Инструкция")
    _write_instruction(
        wi,
        [
            ("ШАБЛОН ПРОДАЖ — ИНСТРУКЦИЯ", True, 16),
            ("", False, 11),
            ("1. Куда вносить данные", True, 13),
            (
                "Только на лист «Данные». Строка 1 — заголовки, со строки 2 — продажи. "
                "Лист «Инструкция» модуль при загрузке игнорирует — можете его читать спокойно.",
                False,
                11,
            ),
            ("", False, 11),
            ("2. Обязательные колонки", True, 13),
            ("Артикул — код товара.", False, 11),
            ("Наименование — название товара.", False, 11),
            ("Дата — дата продажи (ДД.ММ.ГГГГ).", False, 11),
            ("Количество — сколько продано.", False, 11),
            ("", False, 11),
            ("3. Необязательные колонки", True, 13),
            ("Сумма — выручка. Если заполнить, ABC пойдёт по сумме; если пусто — по количеству.", False, 11),
            ("", False, 11),
            ("4. Как заполнять", True, 13),
            ("Одна строка = один товар в одну дату (или один документ).", False, 11),
            ("Один товар в разные дни — несколько строк с одним артикулом.", False, 11),
            ("Период в модуле указывайте по датам, которые есть в этом файле.", False, 11),
            ("Наименования не сокращайте.", False, 11),
            ("", False, 11),
            ("5. Чего нельзя делать", True, 13),
            ("Не переименовывайте лист «Данные».", False, 11),
            ("Не меняйте заголовки строки 1 без необходимости.", False, 11),
            ("Не пишите инструкцию и пояснения на листе «Данные» выше таблицы.", False, 11),
            ("Не оставляйте пустую дату в строках с продажами.", False, 11),
            ("", False, 11),
            ("6. Как загрузить в модуль", True, 13),
            ("Сохраните файл → в программе шаг «Загрузите файл продаж» → выберите этот файл.", False, 11),
            ("", False, 11),
            ("7. Примеры", True, 13),
            ("Жёлтые строки — примеры. Перед реальной работой замените своими данными из 1С.", False, 11),
        ],
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def main() -> None:
    TEMPLATES.mkdir(parents=True, exist_ok=True)
    stock = create_stock_template(TEMPLATES / "Шаблон_остатки.xlsx")
    sales = create_sales_template(TEMPLATES / "Шаблон_продажи.xlsx")
    print(stock)
    print(sales)


if __name__ == "__main__":
    main()
