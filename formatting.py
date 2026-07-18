"""Стили и условное форматирование openpyxl."""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config.settings import SETTINGS


THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

FILL_HEADER = PatternFill("solid", fgColor=SETTINGS["color_header"])
FONT_HEADER = Font(name="Calibri", bold=True, color=SETTINGS["color_header_font"], size=11)
FILL_EDIT = PatternFill("solid", fgColor=SETTINGS["color_edit"])
FILL_FORMULA = PatternFill("solid", fgColor=SETTINGS["color_formula"])
FILL_READONLY = PatternFill("solid", fgColor=SETTINGS["color_readonly"])
FILL_A = PatternFill("solid", fgColor=SETTINGS["color_a"])
FILL_B = PatternFill("solid", fgColor=SETTINGS["color_b"])
FILL_C = PatternFill("solid", fgColor=SETTINGS["color_c"])
FILL_CRIT = PatternFill("solid", fgColor=SETTINGS["color_critical"])
FILL_WARN = PatternFill("solid", fgColor=SETTINGS["color_warning"])
FILL_OK = PatternFill("solid", fgColor=SETTINGS["color_ok"])
FILL_INFO = PatternFill("solid", fgColor=SETTINGS["color_info"])

ALIGN_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
ALIGN_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
ALIGN_RIGHT = Alignment(vertical="center", horizontal="right")


def style_header_row(ws, row: int = 1, max_col: int = 1) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN


def autosize_columns(ws, min_width: int = 10, max_width: int = 55, name_col: int | None = None) -> None:
    """Подбор ширины; для колонки наименования — шире."""
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        letter = get_column_letter(col_idx)
        if name_col and col_idx == name_col:
            ws.column_dimensions[letter].width = max_width
            continue
        length = 0
        for cell in col_cells[:80]:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(val), max_width))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def add_abc_validation(ws, cell_range: str) -> None:
    dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=True)
    dv.error = "Выберите A, B или C"
    dv.errorTitle = "ABC"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_oos_conditional(ws, status_col_letter: str, start_row: int, end_row: int) -> None:
    """Подсветка критичных статусов."""
    rng = f"{status_col_letter}{start_row}:{status_col_letter}{end_row}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Критический",{status_col_letter}{start_row}))'],
            fill=FILL_CRIT,
        ),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("out-of-stock",{status_col_letter}{start_row}))'],
            fill=FILL_WARN,
        ),
    )


def add_order_highlight(ws, order_col: str, start_row: int, end_row: int) -> None:
    rng = f"{order_col}{start_row}:{order_col}{end_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_OK),
    )
